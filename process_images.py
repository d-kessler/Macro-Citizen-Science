import csv
import io
import json
import os
import random
import shutil
import statistics
from datetime import datetime, date

import cv2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from PIL import Image, ImageDraw, ExifTags
from git import Repo
from imutils import perspective, contours, \
    grab_contours
from panoptes_client import Panoptes, Project, \
    SubjectSet, Workflow


class ProcessImages:
    # Predefined folder and file locations
    main_folder = "images"
    # TODO: CHANGE TO JUST "manifests"
    manifests_folder = "manifests - Copy"
    exp_manifest_path = os.path.join(manifests_folder, "Experiment_Manifest.xlsx")
    sim_manifest_path = os.path.join(manifests_folder, "Simulation_Manifest.xlsx")
    neg_manifest_path = os.path.join(manifests_folder, "Negative_Manifest.xlsx")
    sim_feedback_id = "meltpatch"
    neg_feedback_id = "no_meltpatch"
    sim_template_path = r"sim_tools/sim_template.png"
    # Manifest and .csv file names and fieldnames
    exp_csv_file_name = 'experiment_subjects.csv'
    sim_csv_file_name = 'simulation_subjects.csv'
    neg_csv_file_name = 'negative_subjects.csv'
    exp_metadata_fields = ['!subject_id', '#file_name', '#subfolder', "#parent_folder",
                           '#original_file_name', '#warehouse', '#location', '#granite_type',
                           '#slab_id', '#date_time', '#latitude_longitude', '#columns_or_rows',
                           '#image_dimensions(in)', '#glare_area(mm^2)', '#examinable_area(mm^2)',
                           '#grain_density', '#grain_stats(mm)(number, mean_size,' + \
                           'median_size, size_25_percentile, size_75_percentile)']
    sim_metadata_fields = ['!subject_id', '#file_name', '#training_subject', '#feedback_1_id',
                           '#feedback_1_x', '#feedback_1_y', '#feedback_1_toleranceA',
                           '#feedback_1_toleranceB', '#feedback_1_theta', '#major_to_minor_ratio']
    neg_metadata_fields = ['!subject_id', '#file_name', '#training_subject', '#feedback_1_id']
    # Training probabilities in the form:
    # [probability] * number of images for which probability is applied
    training_chances = [0.5] * 4 + [0.4] * 50 + [0.2] * 50
    training_default = [0.10]
    # Ideal final image dimensions (height, width) in inches; allowed to vary (see run())
    image_dimensions = [6, 8]
    # Conversion to millimeters
    image_dimensions_mm = [dim * 25.4 for dim in image_dimensions]
    # Variable deciding whether images are cropped into four parts; allowed to vary ^
    should_crop_into_four = 'y'
    # Number of training images (simulations and negatives each) made per folder
    training_images_per_folder = 5
    # Buffer around the images' edges where (the centers of) simulations will not be drawn
    edges_buffer_mm = 3
    # TODO: TEMPORARY
    lower_limit = 2

    def __init__(self, username, password, project_id, workflow_id, upload_now='n', experiment_set_id=95868,
                 simulation_set_id=95869, negative_set_id=95870):

        # Initializing Zooniverse info (username, password, project & workflow IDs)
        self.project_info = (username, password, project_id, workflow_id)

        # Initializing subject sets
        self.experiment_set_id = experiment_set_id
        self.simulation_set_id = simulation_set_id
        self.negative_set_id = negative_set_id

        # Deciding whether to upload images post-processing
        self.upload_now = upload_now

        # TODO: TEMPORARY
        self.should_draw_scale_bars = input('Should scale bars be drawn on the images? [y/n]: ')
        while self.should_draw_scale_bars != 'y' and self.should_draw_scale_bars != 'n':
            print('Please enter \'y\' or \'no\'... ')
            self.should_draw_scale_bars = input('Should scale bars be drawn on the images? [y/n]: ')

        # Getting a list of parent folders in the main folder
        self.parent_folders = [f.name for f in os.scandir(self.main_folder) if f.is_dir()]

    def push_manifests(self):
        # Pushing the updated local manifests to Github
        repo_directory = '.'  # ie. current directory
        repo = Repo(repo_directory)
        files_to_push = [self.exp_manifest_path, self.sim_manifest_path, self.neg_manifest_path]
        commit_message = f"update manifests, date: {date.today()}"
        repo.index.add(files_to_push)
        repo.index.commit(commit_message)
        origin = repo.remote('origin')
        origin.push(force=True)
        print('\nManifests pushed.')

    def upload_subjects(self, exp_csv_file_path, sim_csv_file_path, neg_csv_file_path):
        exp_upload_cmd = f"panoptes subject-set upload-subjects {self.experiment_set_id} {exp_csv_file_path}"
        sim_upload_cmd = f"panoptes subject-set upload-subjects {self.simulation_set_id} {sim_csv_file_path}"
        neg_upload_cmd = f"panoptes subject-set upload-subjects {self.negative_set_id} {neg_csv_file_path}"
        print("")
        try:
            os.system(exp_upload_cmd)
            print('Experiment subjects uploaded.')
        except:
            print('Error uploading EXPERIMENT subjects; upload manually.')
        try:
            os.system(sim_upload_cmd)
            print('Simulation subjects uploaded.')
        except:
            print('Error uploading SIMULATION subjects; upload manually.')
        try:
            os.system(neg_upload_cmd)
            print('Negative subjects uploaded.')
        except:
            print('Error uploading NEGATIVE subjects; upload manually.')

    def create_negative(self, neg_image_file_path, image_file_name, pos_folder_path):
        pil_img, image_exif = self.configure_pil_img(neg_image_file_path)
        # Displaying the image
        plt.imshow(pil_img)
        plt.show(block=False)
        plt.waitforbuttonpress(0)
        plt.close()
        # Getting input for whether the image is "negative" (contains no candidate melt patches)
        contains_melt_patch = input('Does this image contain a potential melt patch? [y/n]: ')
        while contains_melt_patch != 'y' and contains_melt_patch != 'n':
            print('Please enter \'y\' or \'no\'... ')
            contains_melt_patch = input('Does this image contain a potential melt patch? [y/n]: ')
        # Naming, saving the image according to the above input
        if contains_melt_patch == 'n':
            classification = 'negative'
            pil_img.save(neg_image_file_path, exif=image_exif)
        else:
            classification = 'positive'
            pos_file_name = r"pos_" + image_file_name
            pos_file_path = os.path.join(pos_folder_path, pos_file_name)
            pil_img.save(pos_file_path, exif=image_exif)
            print('\nThe image has been saved to the "positives folder."\n')
        return contains_melt_patch, classification

    @staticmethod
    def create_blank_pil_img(model_pil_img):
        # Creating a blank (white) image of equal dimensions to model_pil_img, saving to blank_img_path
        blank_img_path = r"sim_tools\blank.png"
        blank_img_np = 255 * np.ones(np.array(model_pil_img).shape, np.uint8)
        cv2.imwrite(blank_img_path, blank_img_np)
        # Creating a PIL instance for the blank image
        blank_pil_img = Image.open(blank_img_path)
        return blank_pil_img

    @staticmethod
    def get_glare_ellipse_overlap(glare_pixels, ellipse_pil_img, ellipse_center_coordinates, blank_pil_img):
        # Pasting the simulation ellipse onto a blank (white) image to obtain a list of ellipse pixels
        blank_pil_img.paste(ellipse_pil_img, ellipse_center_coordinates, ellipse_pil_img)
        ellipse_img = np.array(blank_pil_img)
        ellipse_rows, ellipse_cols, _ = np.where(ellipse_img != 255)
        ellipse_pixels = [(row, col) for (row, col) in zip(ellipse_rows, ellipse_cols)]
        overlap = set(glare_pixels).intersection(ellipse_pixels)
        return overlap

    def draw_sim(self, sim_image_file_path):
        # Creating a PIL instance for the present image and simulation ellipse template
        sim_pil_img, sim_image_exif = self.configure_pil_img(sim_image_file_path)
        ellipse_pil_img = Image.open(self.sim_template_path)
        # Getting  list of the image's glare pixels
        _, glare_pixels = self.get_glare_area(sim_image_file_path)
        # Getting the image's millimeter per pixel ratio
        mm_per_pixel = self.get_mm_per_pixel(pil_img=sim_pil_img)
        # Initializing ellipse axes lengths (mm) , and angle (in degrees) rotated above the horizontal
        # TODO: select according to distribution
        minor_axis_mm = 3
        # TODO: select according to distribution
        major_to_minor_axes_ratio = 2
        major_axis_mm = minor_axis_mm * major_to_minor_axes_ratio
        ellipse_axes_lengths = (major_axis_mm, minor_axis_mm)
        # Initializing (random) ellipse position, excluding region around the image's edges
        ellipse_center_x_coord = random.randint(int(self.edges_buffer_mm / mm_per_pixel),
                                                int((self.image_dimensions_mm[
                                                         1] - self.edges_buffer_mm) / mm_per_pixel))
        ellipse_center_y_coord = random.randint(int(self.edges_buffer_mm / mm_per_pixel),
                                                int((self.image_dimensions_mm[
                                                         0] - self.edges_buffer_mm) / mm_per_pixel))
        ellipse_center_coordinates = (ellipse_center_x_coord, ellipse_center_y_coord)
        # Initializing (random) ellipse angle (in degrees) above the horizontal
        ellipse_angle = random.randint(0, 180)
        # Converting axes lengths from millimeters to pixels, accounting for the template image's blank-space borders
        major_axis_pix = major_axis_mm / mm_per_pixel
        minor_axis_pix = minor_axis_mm / mm_per_pixel
        major_axis_excess_pix = 28 * (major_axis_pix / 124)
        minor_axis_excess_pix = 28 * (minor_axis_pix / 124)
        adjusted_major_axis_pix = int(major_axis_pix + major_axis_excess_pix)
        adjusted_minor_axis_pix = int(minor_axis_pix + minor_axis_excess_pix)
        adjusted_ellipse_axes_lengths = (adjusted_major_axis_pix, adjusted_minor_axis_pix)
        # Resizing the ellipse
        ellipse_pil_img = ellipse_pil_img.resize(adjusted_ellipse_axes_lengths)  # major, minor
        # Correcting for the errors introduced by blank-space borders / rotation on the ellipse center coordinates
        M, m = ellipse_pil_img.size
        beta = np.arccos(1 / np.sqrt((m ** 2 / M ** 2) + 1))
        d = np.sqrt((1 / 4) * m ** 2 + (1 / 4) * M ** 2)
        topLeft_to_center_x, topLeft_to_center_y = (None, None)
        if 0 < ellipse_angle <= 90:
            alpha = np.radians(ellipse_angle) - beta
            x1 = d * np.cos(alpha)
            y1 = d * np.sin(alpha)
            b = M * np.sin(np.radians(ellipse_angle))
            x0 = abs(x1)
            y0 = abs(b - y1)
            topLeft_to_center_x = int(ellipse_center_x_coord - x0)
            topLeft_to_center_y = int(ellipse_center_y_coord - y0)
        elif 90 < ellipse_angle <= 180:
            alpha = np.radians(ellipse_angle) + beta
            x1 = d * np.cos(alpha)
            y1 = d * np.sin(alpha)
            u = m * np.sin(np.radians(ellipse_angle - 90))
            x0 = abs(x1)
            y0 = abs(u + y1)
            topLeft_to_center_x = int(ellipse_center_x_coord - x0)
            topLeft_to_center_y = int(ellipse_center_y_coord - y0)
        elif ellipse_angle == 0:
            topLeft_to_center_x = int(ellipse_center_x_coord - (adjusted_major_axis_pix / 2))
            topLeft_to_center_y = int(ellipse_center_y_coord - (adjusted_minor_axis_pix / 2))
        # Rotating the ellipse
        ellipse_pil_img = ellipse_pil_img.rotate(ellipse_angle, expand=True)
        # Ensuring that the ellipse will not be pasted over glare areas
        overlap = 0
        while overlap != set():
            # Pasting the simulation ellipse onto a blank (white) image to obtain a list of ellipse pixels
            blank_pil_img = self.create_blank_pil_img(sim_pil_img)
            overlap = self.get_glare_ellipse_overlap(glare_pixels, ellipse_pil_img, ellipse_center_coordinates,
                                                     blank_pil_img)
        # Pasting the ellipse onto pil_image, saving to the given file path
        sim_pil_img.paste(ellipse_pil_img, (topLeft_to_center_x, topLeft_to_center_y), ellipse_pil_img)
        sim_pil_img.save(sim_image_file_path, exif=sim_image_exif)
        return ellipse_axes_lengths, ellipse_center_coordinates, ellipse_angle, major_to_minor_axes_ratio

    @staticmethod
    def write_metadata_into_CSV(CSV_file_path, metadata_fields, metadata_list):
        with open(CSV_file_path, 'a', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)
            row = {}
            for i in range(len(metadata_fields)):
                row[str(metadata_fields[i])] = metadata_list[i]
            csv_writer.writerow(row)

    @staticmethod
    def write_metadata_into_manifest(ws, row, metadata_list):
        for i in range(len(metadata_list)):
            ws.cell(row=row, column=(i + 1)).value = str(metadata_list[i])

    @staticmethod
    def convert_to_degrees(value):
        # Converting GPS exif data to latitude/longitude degree format
        d0 = value[0][0]
        d1 = value[0][1]
        d = float(d0) / float(d1)
        m0 = value[1][0]
        m1 = value[1][1]
        m = float(m0) / float(m1)
        s0 = value[2][0]
        s1 = value[2][1]
        s = float(s0) / float(s1)

        return d + (m / 60.0) + (s / 3600.0)

    def get_gps_exif(self, exif_dict):
        gps_dict = {}
        try:
            for key in exif_dict['GPSInfo'].keys():
                gps_tag = ExifTags.GPSTAGS.get(key)
                gps_dict[gps_tag] = exif_dict['GPSInfo'][key]
            latitude_raw = gps_dict.get('GPSLatitude')
            longitude_raw = gps_dict.get('GPSLongitude')
            lat_ref = gps_dict.get('GPSLatitudeRef')
            long_ref = gps_dict.get('GPSLongitudeRef')
            if lat_ref == "S":
                latitude = -abs(self.convert_to_degrees(latitude_raw))
            else:
                latitude = self.convert_to_degrees(latitude_raw)
            if long_ref == "W":
                longitude = -abs(self.convert_to_degrees(longitude_raw))
            else:
                longitude = self.convert_to_degrees(longitude_raw)
        except:
            latitude = None
            longitude = None
            pass
        latitude_longitude = [latitude, longitude]
        return latitude_longitude

    @staticmethod
    def get_date_exif(exif_dict):
        # Attempting to retrieve data exif data
        try:
            date_exif = datetime.strptime(exif_dict['DateTimeOriginal'], '%Y:%m:%d %H:%M:%S')
            date_exif = json.dumps(date_exif, default=str)
        except:
            date_exif = None
            pass
        return date_exif

    def resize_to_limit(self, image_file_path, size_limit=600000):
        pil_img, image_exif = self.configure_pil_img(image_file_path)
        aspect = pil_img.size[0] / pil_img.size[1]
        while True:
            with io.BytesIO() as buffer:
                pil_img.save(buffer, format="JPEG")
                data = buffer.getvalue()
            file_size = len(data)
            size_deviation = file_size / size_limit
            if size_deviation <= 1:
                pil_img.save(image_file_path, exif=image_exif)
                break
            else:
                new_width = pil_img.size[0] / (size_deviation ** 0.5)
                new_height = new_width / aspect
                pil_img = pil_img.resize((int(new_width), int(new_height)))

    def draw_scale_bars(self, image_file_path, thickness):
        pil_img, image_exif = self.configure_pil_img(image_file_path)
        mm_per_pixel = self.get_mm_per_pixel(pil_img=pil_img)
        pix_width, pix_height = pil_img.size
        draw = ImageDraw.Draw(pil_img)
        # Setting scale bar color, length, and number
        scale_bars_color = (100, 255, 0)  # (R, G, B)
        scale_bar_pix_length = self.lower_limit / mm_per_pixel
        number_of_scale_bars = 10
        # Configuring left/right side scale bars
        leftRight_bars_center = (pix_height / number_of_scale_bars) / 2
        leftRight_start_coords = [15, leftRight_bars_center - (.5 * scale_bar_pix_length)]
        leftRight_end_coords = [leftRight_start_coords[0], leftRight_start_coords[1] + scale_bar_pix_length]
        leftRight_coords = [tuple(leftRight_start_coords), tuple(leftRight_end_coords)]
        # Configuring top/bottom side scale bars
        topBottom_bars_center = (pix_width / number_of_scale_bars) / 2
        topBottom_start_coords = [topBottom_bars_center - (.5 * scale_bar_pix_length), 15]
        topBottom_end_coords = [topBottom_start_coords[0] + scale_bar_pix_length, topBottom_start_coords[1]]
        topBottom_coords = [tuple(topBottom_start_coords), tuple(topBottom_end_coords)]
        # Drawing scale bars
        for j in range(1, number_of_scale_bars + 1):
            draw.line(leftRight_coords, fill=scale_bars_color, width=thickness)
            draw.line(topBottom_coords, fill=scale_bars_color, width=thickness)
            leftRight_start_coords[1] += (2 * leftRight_bars_center)
            topBottom_start_coords[0] += (2 * topBottom_bars_center)
            if j % 2 != 0:
                leftRight_start_coords[0] = pix_width - leftRight_start_coords[0]
                topBottom_start_coords[1] = pix_height - topBottom_start_coords[1]
            else:
                leftRight_start_coords[0] = 15
                topBottom_start_coords[1] = 15
            leftRight_end_coords = [leftRight_start_coords[0], leftRight_start_coords[1] + scale_bar_pix_length]
            topBottom_end_coords = [topBottom_start_coords[0] + scale_bar_pix_length, topBottom_start_coords[1]]
            leftRight_coords = [tuple(leftRight_start_coords), tuple(leftRight_end_coords)]
            topBottom_coords = [tuple(topBottom_start_coords), tuple(topBottom_end_coords)]
        # Saving image to image_file_path with its exif data
        pil_img.save(image_file_path, exif=image_exif)

    @staticmethod
    def midpoint(ptA, ptB):
        return (ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5

    @staticmethod
    def canny_thresholds(img, sigma=0.33):
        # Compute the median of the single channel pixel intensities
        v = np.median(img)
        # Apply automatic Canny edge detection using the computed median
        lower = int(max(0, (1.0 - sigma) * v))
        upper = int(min(255, (1.0 + sigma) * v))
        return lower, upper

    def get_contours(self, prepared_cv2_img, cutoff_avg_dim_mm):
        mm_per_pixel = self.get_mm_per_pixel(cv2_img=prepared_cv2_img)
        # Detecting contours
        lower_canny, upper_canny = self.canny_thresholds(prepared_cv2_img)
        edges_img = cv2.Canny(prepared_cv2_img, lower_canny, upper_canny)
        # Dilating to ensure that contour paths are continuous
        dil_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        dil_edges_img = cv2.dilate(edges_img, dil_kernel)
        # Getting, sorting a list of contours
        contours_list = cv2.findContours(dil_edges_img.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours_list = grab_contours(contours_list)
        if contours_list:
            (contours_list, _) = contours.sort_contours(contours_list)
        # Initializing lists
        wanted_contours_list = []
        contour_dims = []
        center_xs = []
        center_ys = []
        for cont in contours_list:
            #  Creating bounding box
            box = cv2.minAreaRect(cont)
            box = cv2.boxPoints(box)
            box = np.array(box, dtype='int')
            box = perspective.order_points(box)
            # Getting box vertex coordinates
            (top_left, top_right, bot_right, bot_left) = box
            # Getting box side midpoints
            (top_mid_x, top_mid_y) = self.midpoint(top_left, top_right)
            (bot_mid_x, bot_mid_y) = self.midpoint(bot_left, bot_right)
            (left_mid_x, left_mid_y) = self.midpoint(top_left, bot_left)
            (right_mid_x, right_mid_y) = self.midpoint(top_right, bot_right)
            # Getting box center coordinates
            center_x = np.average(box[:, 0])
            center_y = np.average(box[:, 1])
            center_xs.append(center_x)
            center_ys.append(center_y)
            # Measuring the semi-major and semi-minor axes of the box in pixels
            pix_height = np.linalg.norm(np.array([top_mid_x, top_mid_y]) - np.array([bot_mid_x, bot_mid_y]))
            pix_width = np.linalg.norm(np.array([left_mid_x, left_mid_y]) - np.array([right_mid_x, right_mid_y]))
            # pix_height = distance.euclidean((top_mid_x, top_mid_y), (bot_mid_x, bot_mid_y))
            # pix_width = distance.euclidean((left_mid_x, left_mid_y), (right_mid_x, right_mid_y))
            contour_dim = (pix_height, pix_width)
            # Converting axes lengths to millimeters
            mm_height = pix_height * mm_per_pixel
            mm_width = pix_width * mm_per_pixel
            # Getting the contour's average dimension (average of height/width) in millimeters
            contour_avg_dim = (mm_height + mm_width) / 2
            # Ignoring contours that have an average dimension less than the specified cutoff
            if contour_avg_dim < cutoff_avg_dim_mm:
                continue
            wanted_contours_list.append(cont)
            contour_dims.append(contour_dim)
        # Getting a list of contour areas in square pixels
        contour_areas = []
        for cont in wanted_contours_list:
            area = cv2.contourArea(cont)
            contour_areas.append(area)
        # Finding the total contour area in square pixels
        total_contour_area = sum(contour_areas)
        return wanted_contours_list, contour_dims, total_contour_area

    def get_glare_area(self, image_file_path):
        cv2_img = cv2.imread(image_file_path)
        mm_per_pixel = self.get_mm_per_pixel(cv2_img=cv2_img)
        # Setting lower and upper color limits for thresholding
        lower = (230, 230, 230)
        upper = (255, 255, 255)
        # Getting pixels between lower and upper color limits (those corresponding to glare areas)
        mask = cv2.inRange(cv2_img, lower, upper)
        # Blurring to smooth edges
        mask = cv2.blur(mask, (10, 10))
        # Getting a list of the glare patches' edges/contours, dimensions and the total glare area
        glare_contours, glare_dims, total_glare_area = self.get_contours(mask, 2.5)
        # Converting the total glare area from square pixels to square millimeters
        total_glare_area_mmSq = total_glare_area * (mm_per_pixel ** 2)
        # Isolating glare areas on a blank (black) image to get a list of glare pixels
        blank_img = np.zeros(cv2_img.shape, np.uint8)
        isolated_glare_img = cv2.drawContours(blank_img, glare_contours, contourIdx=-1, color=(255, 255, 255),
                                              thickness=cv2.FILLED)
        glare_rows, glare_cols, _ = np.where(isolated_glare_img != 0)
        glare_pixels = [(row, col) for (row, col) in zip(glare_rows, glare_cols)]
        return total_glare_area_mmSq, glare_pixels

    def get_grain_stats(self, image_file_path):
        """Note: "grains" refers to the dark-colored grains in the granite"""
        # Configuring the original image, getting its 'millimeter per pixel' scale
        cv2_img = cv2.imread(image_file_path)
        mm_per_pixel = self.get_mm_per_pixel(cv2_img=cv2_img)
        # Converting to grayscale
        gray_img = cv2.cvtColor(cv2_img, cv2.COLOR_BGR2GRAY)
        # Blurring grayscale image
        blurred_img = cv2.GaussianBlur(gray_img, (9, 9), 0)
        # Segmenting blurred image to separate dark grains from background
        segmentation_thresh = 70
        thresh_img = cv2.threshold(blurred_img, segmentation_thresh, 255, cv2.THRESH_BINARY)[1]
        # Opening (erosion followed by dilation) segmented image to close holes within grains
        kernel = np.ones((9, 9), np.uint8)
        opened_img = cv2.morphologyEx(thresh_img, cv2.MORPH_OPEN, kernel, 5)
        # Getting a list of the grains' edges/contours, dimensions, and total area
        grain_contours_list, grain_dims, total_grain_area = self.get_contours(opened_img, 0.015)
        # Converting the grain dimensions from pixels to millimeters
        grain_dims_mm = [(x * mm_per_pixel, y * mm_per_pixel) for x, y in grain_dims]
        # Finding the total image area in square pixels
        image_area_pix = (self.image_dimensions_mm[0] * self.image_dimensions_mm[1]) / (mm_per_pixel ** 2)
        # Getting the image's grain density (area of grains / area of image)
        grain_density = total_grain_area / image_area_pix
        # Getting statistics on the image's grains
        h = [g[0] for g in grain_dims_mm]
        w = [g[1] for g in grain_dims_mm]
        number_of_grains = len(grain_dims_mm)
        mean_grain_size = statistics.mean([statistics.mean(h), statistics.mean(w)])
        median_grain_size = statistics.median([statistics.median(h), statistics.median(w)])
        grain_size_25_percentile = statistics.mean([np.percentile(h, 25), np.percentile(w, 25)])
        grain_size_75_percentile = statistics.mean([np.percentile(h, 75), np.percentile(w, 75)])
        grain_stats = (
            number_of_grains, mean_grain_size, median_grain_size, grain_size_25_percentile, grain_size_75_percentile)
        return grain_density, grain_stats

    def get_mm_per_pixel(self, cv2_img=None, pil_img=None, image_file_path=None):
        pix_height = None
        if image_file_path:
            cv2_img = cv2.imread(image_file_path)
        if cv2_img is not None:
            try:
                pix_height, _, _ = cv2_img.shape
            except ValueError:
                pix_height, _ = cv2_img.shape
        elif pil_img:
            _, pix_height = pil_img.size
        mm_per_pixel = self.image_dimensions_mm[0] / pix_height
        return mm_per_pixel

    def assign_image_name(self, image_file_name, exp_subject_id, slab_id, row, col, warehouse, location):
        extension = os.path.splitext(image_file_name)[-1]
        if self.should_crop_into_four == 'y':
            crop_file_name = os.path.splitext(image_file_name)[-2]
            crop_location = crop_file_name.split("_")[-1]
            exp_file_name = exp_subject_id + r"_" + str(slab_id) + r"_" + str(row) + r"_" + str(col) + \
                            r"_" + crop_location + r"_" + warehouse + r"_" + location + r"_" + str(extension)
        else:
            exp_file_name = exp_subject_id + r"_" + str(slab_id) + r"_" + str(row) + r"_" + str(col) + \
                            r"_" + warehouse + r"_" + location + r"_" + str(extension)
        return exp_file_name

    def assign_image_position(self, image_file_names, image_file_name, imaging_scheme,
                              number_of_columns_or_rows, row, col):
        image_index = image_file_names.index(image_file_name) + 1
        number_of_columns_or_rows = int(number_of_columns_or_rows)
        if self.should_crop_into_four == 'y':
            if image_index % 4 == 0 and (image_index / 4) % number_of_columns_or_rows != 0:
                if imaging_scheme == 'columns':
                    if row % 2 == 1:
                        col += 1
                    elif row % 2 == 0:
                        col -= 1
                elif imaging_scheme == 'rows':
                    if col % 2 == 1:
                        row += 1
                    elif col % 2 == 0:
                        row -= 1
            elif image_index % 4 == 0 and (image_index / 4) % number_of_columns_or_rows == 0:
                if imaging_scheme == 'columns':
                    row += 1
                elif imaging_scheme == 'rows':
                    col += 1
        elif self.should_crop_into_four == 'n':
            if image_index % number_of_columns_or_rows != 0:
                if imaging_scheme == 'columns':
                    if row % 2 == 1:
                        col += 1
                    elif row % 2 == 0:
                        col -= 1
                elif imaging_scheme == 'rows':
                    if col % 2 == 1:
                        row += 1
                    elif col % 2 == 0:
                        row -= 1
            elif image_index % number_of_columns_or_rows == 0:
                if imaging_scheme == 'columns':
                    row += 1
                elif imaging_scheme == 'rows':
                    col += 1
        return row, col

    @staticmethod
    def configure_pil_img(image_file_path):
        pil_img = Image.open(image_file_path)
        try:
            image_exif = pil_img.info['exif']
        except KeyError:
            image_exif = pil_img.getexif()
        return pil_img, image_exif

    @staticmethod
    def get_file_names(folder_path):
        all_file_names = os.listdir(folder_path)
        wanted_file_names = []
        for file_name in all_file_names:
            if file_name.endswith(".jpeg") or file_name.endswith(".jpg") or file_name.endswith(".png"):
                wanted_file_names.append(file_name)
        return wanted_file_names

    @staticmethod
    def configure_exp_metadata(parent_folder, subfolder):
        exp_metadata = {}
        date, _, warehouse, city, state = parent_folder.split("_")
        location = city + ", " + state
        slab_id, granite_type, columns_or_rows = subfolder.split("_")
        if '|' or '/' or '\\' in slab_id:
            slab_id = slab_id.replace('|', '-')
            slab_id = slab_id.replace('/', '-')
            slab_id = slab_id.replace('\\', '-')
        imaging_scheme, number_of_columns_or_rows = (None, None)
        if 'c' in columns_or_rows:
            imaging_scheme = 'columns'
            number_of_columns_or_rows = columns_or_rows.split('c')[-1]
        elif 'r' in columns_or_rows:
            imaging_scheme = 'rows'
            number_of_columns_or_rows = columns_or_rows.split('r')[-1]
        exp_metadata['date'] = date
        exp_metadata['warehouse'] = warehouse
        exp_metadata['location'] = location
        exp_metadata['granite_type'] = granite_type
        exp_metadata['slab_id'] = slab_id
        exp_metadata['imaging_scheme'] = imaging_scheme
        exp_metadata['number_of_columns_or_rows'] = number_of_columns_or_rows
        return exp_metadata

    @staticmethod
    def configure_CSV(subfolder_path, desired_csv_file_name, metadata_fields_list=None):
        if metadata_fields_list is None:
            metadata_fields_list = ['']
        desired_csv_file_path = os.path.join(subfolder_path, desired_csv_file_name)
        with open(desired_csv_file_path, 'w', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields_list)
            csv_writer.writeheader()
        return desired_csv_file_path

    @staticmethod
    def configure_excel(excel_file_path):
        wb = openpyxl.load_workbook(filename=excel_file_path)
        ws = wb['Sheet1']
        first_empty_row = None
        for row in range(1, int(1e10)):
            if ws.cell(row, 1).value is None:
                first_empty_row = row
                break
        return wb, ws, first_empty_row

    def create_negative_images(self, exp_folder_path, neg_folder_path, pos_folder_path):
        # Configuring negative manifest
        neg_wb, neg_ws, neg_i = self.configure_excel(self.neg_manifest_path)
        # Configuring negative csv file
        neg_csv_file_path = self.configure_CSV(neg_folder_path, self.neg_csv_file_name,
                                               self.neg_metadata_fields)
        image_file_names_sample_negs = random.sample(self.get_file_names(exp_folder_path),
                                                     self.training_images_per_folder)
        for image_file_name in image_file_names_sample_negs:
            # Getting the image's file path
            image_file_path = os.path.join(exp_folder_path, image_file_name)
            # Assigning a negative subject ID
            neg_subject_id = 'n' + str(neg_i - 1)
            # Renaming, moving the image to the negative folder, creating a Pillow image instance, and getting exif info
            neg_image_file_name = neg_subject_id + r"_" + image_file_name
            neg_image_file_path = os.path.join(neg_folder_path, neg_image_file_name)
            neg_image_file_path = shutil.copyfile(image_file_path, neg_image_file_path)
            # Getting input for whether the image is "negative" (contains no candidate melt patches)
            if image_file_names_sample_negs.index(image_file_name) == 0:
                print("\n\nMaking negative images...")
            print('Displaying image {} of {}'.format(image_file_names_sample_negs.index(image_file_name) + 1,
                                                     len(image_file_names_sample_negs)))
            contains_melt_patch, classification = self.create_negative(neg_image_file_path, image_file_name,
                                                                       pos_folder_path)
            if contains_melt_patch == 'y':
                continue
            # Resizing the image to the Zooniverse recommended 600MB and saving changes
            self.resize_to_limit(image_file_path=neg_image_file_path)
            # Writing the image's metadata into the negative manifest and negative CSV
            neg_metadata_list = [neg_subject_id, neg_image_file_name, "True", self.neg_feedback_id]
            self.write_metadata_into_manifest(neg_ws, neg_i, neg_metadata_list)
            self.write_metadata_into_CSV(neg_csv_file_path, self.neg_metadata_fields, neg_metadata_list)
            # Updating indexing variable
            neg_i += 1
        # Saving negative manifest
        neg_wb.save(self.neg_manifest_path)
        return neg_csv_file_path

    def create_simulation_images(self, exp_folder_path, sim_folder_path):
        # Configuring simulation manifest
        sim_wb, sim_ws, sim_i = self.configure_excel(self.sim_manifest_path)
        # Configuring simulation csv file
        sim_csv_file_path = self.configure_CSV(sim_folder_path, self.sim_csv_file_name,
                                               self.sim_metadata_fields)
        image_file_names_sample_sims = random.sample(self.get_file_names(exp_folder_path),
                                                     self.training_images_per_folder)
        for image_file_name in image_file_names_sample_sims:
            # Getting the image's file path
            image_file_path = os.path.join(exp_folder_path, image_file_name)
            # Assigning a simulation subject ID
            sim_subject_id = 's' + str(sim_i - 1)
            # Renaming, moving the image to the simulation folder, creating a Pillow image instance, and getting exif info
            sim_image_file_name = sim_subject_id + r"_" + image_file_name
            sim_image_file_path = os.path.join(sim_folder_path, sim_image_file_name)
            sim_image_file_path = shutil.copyfile(image_file_path, sim_image_file_path)
            # Drawing the simulation
            ellipse_axes_lengths, ellipse_center_coordinates, ellipse_angle, major_to_minor_axes_ratio \
                = self.draw_sim(sim_image_file_path)
            # TODO: TEMPORARY
            # If scale bars were drawn, redrawing them to ensure than they were not covered by a simulation
            if self.should_draw_scale_bars == 'y':
                self.draw_scale_bars(sim_image_file_path, 8)
            # Resizing the image to the Zooniverse recommended 600MB and saving changes
            self.resize_to_limit(sim_image_file_path)
            # Getting images' millimeter per pixel ratio
            mm_per_pixel = self.get_mm_per_pixel(image_file_path=sim_image_file_path)
            # Writing the image's metadata into the simulation manifest and simulation CSV
            sim_metadata_list = [sim_subject_id, sim_image_file_name, "True", self.sim_feedback_id,
                                 ellipse_center_coordinates[0], ellipse_center_coordinates[1],
                                 int(ellipse_axes_lengths[0] / mm_per_pixel),
                                 int(ellipse_axes_lengths[1] / mm_per_pixel),
                                 ellipse_angle, major_to_minor_axes_ratio]
            self.write_metadata_into_manifest(sim_ws, sim_i, sim_metadata_list)
            self.write_metadata_into_CSV(sim_csv_file_path, self.sim_metadata_fields, sim_metadata_list)
            # Formatting command-line interface
            if image_file_names_sample_sims.index(image_file_name) == 0:
                print("")
            # Printing status
            print('\r{} of {} simulations made.'.format(
                (image_file_names_sample_sims.index(image_file_name) + 1),
                len(image_file_names_sample_sims)), end="")
            # Updating indexing variable
            sim_i += 1
        # Saving simulation manifest
        sim_wb.save(self.sim_manifest_path)
        return sim_csv_file_path

    def create_experiment_images(self, parent_folder, subfolder, raw_images_folder_path, exp_folder_path):
        # Configuring experiment images' metadata
        exp_metadata = self.configure_exp_metadata(parent_folder, subfolder)
        # Configuring experiment manifest
        exp_wb, exp_ws, exp_i = self.configure_excel(self.exp_manifest_path)
        # Configuring experiment csv file
        exp_csv_file_path = self.configure_CSV(exp_folder_path, self.exp_csv_file_name,
                                               self.exp_metadata_fields)
        # Initializing variables used to position images on the slab
        row = 1
        col = 0
        image_file_names = self.get_file_names(raw_images_folder_path)
        for image_file_name in image_file_names:
            # Getting the image's file path
            image_file_path = os.path.join(raw_images_folder_path, image_file_name)
            # Assigning an experiment subject ID
            exp_subject_id = 'e' + str(exp_i - 1)
            # Getting the image's position on the slab and naming its file accordingly
            row, col = self.assign_image_position(image_file_names, image_file_name, exp_metadata['imaging_scheme'],
                                                  exp_metadata['number_of_columns_or_rows'], row, col)
            exp_file_name = self.assign_image_name(image_file_name, exp_subject_id, exp_metadata['slab_id'], row, col,
                                                   exp_metadata['warehouse'],
                                                   exp_metadata['location'].replace(", ", "_"))
            # Naming / moving the image to the experiment folder and getting images' exif information
            exp_image_file_path = os.path.join(exp_folder_path, exp_file_name)
            exp_image_file_path = shutil.copyfile(image_file_path, exp_image_file_path)
            # Getting the mean grain size and grain density of the image's granite
            grain_density, grain_stats = self.get_grain_stats(exp_image_file_path)
            # Getting the image's area, total glare area, and examinable area
            image_area_mmSq = self.image_dimensions_mm[0] * self.image_dimensions_mm[1]
            total_glare_area_mmSq, _ = self.get_glare_area(exp_image_file_path)
            buffered_area_mmSq = 2 * self.edges_buffer_mm * (self.image_dimensions_mm[0] + self.image_dimensions_mm[1])
            examinable_area_mmSq = image_area_mmSq - total_glare_area_mmSq - buffered_area_mmSq
            # Resizing the image to the Zooniverse recommended 600MB
            """Note: this is done here so that scale bars are drawn properly; it will no longer be necessary
            when scale bars are no longer necessary"""
            self.resize_to_limit(exp_image_file_path)
            # TODO: TEMPORARY
            # Drawing scale bars on the image (if specified)
            if self.should_draw_scale_bars == 'y':
                self.draw_scale_bars(exp_image_file_path, thickness=10)
            # Resizing the image to the Zooniverse recommended 600MB
            self.resize_to_limit(exp_image_file_path)
            # Getting data and GPS info from the image's exif data
            exp_pil_img, exp_image_exif = self.configure_pil_img(exp_image_file_path)
            exif_dict = {ExifTags.TAGS[k]: j for k, j in exp_pil_img.getexif().items() if k in ExifTags.TAGS}
            date_exif = self.get_date_exif(exif_dict)
            latitude_longitude = self.get_gps_exif(exif_dict)
            # Writing the image's metadata into the experiment manifest and experiment CSV
            exp_metadata_list = [exp_subject_id, str(exp_file_name), str(subfolder), str(parent_folder),
                                 str(image_file_name).replace(".", ""), exp_metadata['warehouse'],
                                 exp_metadata['location'], exp_metadata['granite_type'], exp_metadata['slab_id'],
                                 date_exif.replace("\"", "") if date_exif is not None else exp_metadata['date'],
                                 str(latitude_longitude[0]) + ", " + str(latitude_longitude[1])
                                 if latitude_longitude is not None else "",
                                 exp_metadata['imaging_scheme'] + str(exp_metadata['number_of_columns_or_rows']),
                                 str(self.image_dimensions_mm[0]) + "x" + str(self.image_dimensions_mm[1]),
                                 total_glare_area_mmSq, examinable_area_mmSq, grain_density, grain_stats]
            self.write_metadata_into_manifest(exp_ws, exp_i, exp_metadata_list)
            self.write_metadata_into_CSV(exp_csv_file_path, self.exp_metadata_fields, exp_metadata_list)
            # Printing status
            print('\r{} of {} images processed.'.format((image_file_names.index(image_file_name) + 1),
                                                        len(image_file_names)), end="")
            # Updating indexing variable
            exp_i += 1
        # Saving experiment manifest
        exp_wb.save(self.exp_manifest_path)
        return exp_csv_file_path

    @staticmethod
    def crop_into_four(image_file_names, current_folder_path, cropped_folder_path):
        for file_name in image_file_names:
            image_file_path = os.path.join(current_folder_path, file_name)
            cropped_file_path = os.path.join(cropped_folder_path, file_name)
            extension = os.path.splitext(file_name)[-1]
            original_pil_img = Image.open(image_file_path)
            try:
                image_exif = original_pil_img.info['exif']
            except KeyError:
                image_exif = original_pil_img.getexif()
            width, height = original_pil_img.size
            half_width = width / 2
            half_height = height / 2
            # Starting from top left (0,0) and moving clockwise
            section_1 = (0, 0, half_width, half_height)  # (Left, Top, Right, Bottom) Starting pts.
            section_2 = (half_width, 0, width, half_height)
            section_3 = (half_width, half_height, width, height)
            section_4 = (0, half_height, half_width, height)
            cropped_pil_img = None
            location = ''
            for j in range(1, 5):
                if j == 1:
                    cropped_pil_img = original_pil_img.crop(section_1)
                    location = 'TL'
                if j == 2:
                    cropped_pil_img = original_pil_img.crop(section_2)
                    location = 'TR'
                if j == 3:
                    cropped_pil_img = original_pil_img.crop(section_3)
                    location = 'BR'
                if j == 4:
                    cropped_pil_img = original_pil_img.crop(section_4)
                    location = 'BL'
                reformatted_cropped_file_path = cropped_file_path.replace(str(extension),
                                                                          f"_{location}{str(extension)}")
                cropped_pil_img.save(reformatted_cropped_file_path, exif=image_exif)
        print('\nImages cropped.\n')

    @staticmethod
    def clear_folder(folder_path):
        try:
            shutil.rmtree(folder_path)
            os.mkdir(folder_path)
        except PermissionError or FileNotFoundError:
            input('\nPermissions error for {}, '
                  '\nExit the file window and restart the program.'.format(folder_path))

    def make_folder(self, current_folder_path, desired_folder_name=""):
        desired_folder_path = os.path.join(current_folder_path, desired_folder_name)
        try:
            os.mkdir(desired_folder_path)
        except FileExistsError:
            self.clear_folder(desired_folder_path)
        return desired_folder_path

    def make_folders(self, current_folder_path):
        experiment_folder_path = self.make_folder(current_folder_path, r"experiment")
        # TODO: TEMPORARY(?)
        if self.should_crop_into_four == 'y':
            cropped_folder_path = self.make_folder(current_folder_path, r"cropped")
        else:
            cropped_folder_path = ''
        sim_folder_path = self.make_folder(current_folder_path, r"simulations")
        neg_folder_path = self.make_folder(current_folder_path, r"negatives")
        pos_folder_path = self.make_folder(current_folder_path, r"positives")
        return experiment_folder_path, cropped_folder_path, sim_folder_path, neg_folder_path, pos_folder_path

    def configure_designator(self, username, password, project_id, workflow_id, simulation_set_id, negative_set_id,
                             training_chances, training_default):
        # Connect to project & workflow
        zoonv_project, zoonv_workflow \
            = self.configure_zooniverse(username, password, project_id, workflow_id)
        # Configuring training sets, training image probabilities
        zoonv_workflow.configuration['training_set_ids'] = [simulation_set_id, negative_set_id]
        zoonv_workflow.configuration['training_chances'] = training_chances
        zoonv_workflow.configuration['training_default_chances'] = training_default
        zoonv_workflow.configuration[
            'subject_queue_page_size'] = 10  # determines how many subjects are loaded in queue at one time
        # Training subjects are not retired, experiment subjects are retired via SWAP/Caesar
        zoonv_workflow.retirement['criteria'] = 'never_retire'
        # Change to 15? -- No, Caesar retires experiment subjects without retiring training subjects
        # Saving
        zoonv_workflow.modified_attributes.add('configuration')
        zoonv_workflow.save()
        print('\nDesignator configured.')

    @staticmethod
    def print_existing_subject_sets(subject_set_ids, subject_set_names):
        print('\nThe existing subject sets are:')
        print(
            "\n".join(
                u"{} \u2014 {}".format(ss_id, ss_name) for ss_id, ss_name in zip(subject_set_ids, subject_set_names)))

    def get_existing_subject_sets(self, username, password, project_id, workflow_id):
        # Connecting to project & workflow
        zoonv_project, zoonv_workflow \
            = self.configure_zooniverse(username, password, project_id, workflow_id)
        # Getting and printing a list of existing subject sets' names & IDs
        subject_set_ids = []
        subject_set_names = []
        for ss in zoonv_project.links.subject_sets:
            subject_set_name = ss.display_name
            subject_set_id = int(str(ss).split()[1].replace('>', ''))
            subject_set_ids.append(subject_set_id)
            subject_set_names.append(subject_set_name)
        return subject_set_ids, subject_set_names

    @staticmethod
    def create_subject_set(zoonv_project, zoonv_workflow, subject_set_name):
        subject_set = SubjectSet()
        subject_set.links.project = zoonv_project
        subject_set.display_name = subject_set_name
        subject_set.save()
        zoonv_workflow.links.subject_sets.add(subject_set)
        zoonv_workflow.save()
        subject_set_id = int(str(subject_set).split()[1].replace('>', ''))
        return subject_set_id

    def configure_subject_set(self, username, password, project_id, workflow_id, subject_type, subject_set_ids):
        # Connecting to project & workflow
        zoonv_project, zoonv_workflow \
            = self.configure_zooniverse(username, password, project_id, workflow_id)
        # Selecting an existing subject set or creating a new one based on user input
        if subject_type == 'experiment':
            print("")
        need_new_set = input(f"Would you like to create a new {subject_type.upper()} subject set? [y/n]: ")
        while need_new_set != 'y' and need_new_set != 'n':
            print("Please enter \'y\' or \'no\'...")
            need_new_set = input(f"\nWould you like to create a new {subject_type.upper()} subject set? [y/n]: ")
        if need_new_set == 'n':
            subject_set_id = input("    Enter the ID of the existing set you\'d like to upload to: ")
            while (int(subject_set_id) in subject_set_ids) is False:
                subject_set_id = input("    This ID does not exist; please enter a new one: ")
        else:
            subject_set_name = input("    Enter a name for the new subject set: ")
            subject_set_id = self.create_subject_set(zoonv_project, zoonv_workflow, subject_set_name)
        return subject_set_id

    @staticmethod
    def configure_zooniverse(username, password, project_id, workflow_id):
        # Connect to Zooniverse, get project & workflow ID's
        Panoptes.connect(username=username, password=password)
        zoonv_project = Project.find(project_id)
        zoonv_project.save()
        zoonv_workflow = Workflow.find(workflow_id)
        zoonv_workflow.save()
        return zoonv_project, zoonv_workflow

    def configure_subject_sets(self):
        # Getting existing subject sets
        subject_set_ids, subject_set_names = self.get_existing_subject_sets(*self.project_info)
        # Printing existing subject sets
        self.print_existing_subject_sets(subject_set_ids, subject_set_names)
        # Configuring experiment, simulation, and negative subject sets
        experiment_set_id = self.configure_subject_set(*self.project_info, 'experiment', subject_set_ids)
        simulation_set_id = self.configure_subject_set(*self.project_info, 'simulation', subject_set_ids)
        negative_set_id = self.configure_subject_set(*self.project_info, 'negative', subject_set_ids)
        # Configuring designator for selected/created subject sets (possibly redundant, but not time consuming)
        self.configure_designator(*self.project_info, simulation_set_id, negative_set_id,
                                  self.training_chances, self.training_default)
        return experiment_set_id, simulation_set_id, negative_set_id

    @staticmethod
    def configure_image_dimensions(parent_folder):
        image_dimensions = parent_folder.split("_")[1].split("x")
        image_dimensions = [int(dim) for dim in image_dimensions]
        image_dimensions_mm = [dim * 25.4 for dim in image_dimensions]
        should_crop_into_four = 'y'
        if min(image_dimensions) > 7 and max(image_dimensions) > 9:
            should_crop_into_four = 'y'
        elif min(image_dimensions) <= 7 and max(image_dimensions) <= 9:
            should_crop_into_four = 'n'
        return image_dimensions, image_dimensions_mm, should_crop_into_four

    @staticmethod
    def parent_folder_named_correctly(parent_folder):
        pf_components = parent_folder.split("_")
        pf_named_correctly = (len(pf_components) == 5) and \
                             (len(pf_components[1].split("x")) == 2)
        while pf_named_correctly is False:
            parent_folder = input(f"""
                    {parent_folder} is named incorrectly. 
                    The correct format is:
                    MM-DD-YYYY_Dim1xDim2_WarehouseName_WarehouseCity_WarehouseState
                    Input the correct name: 
                    """)
            pf_components = parent_folder.split("_")
            pf_named_correctly = (len(pf_components) == 5) and \
                                 (len(pf_components[1].split("x")) == 2)

    @staticmethod
    def subfolder_named_correctly(subfolder):
        sf_components = subfolder.split("_")
        sf_named_correctly = (len(sf_components) == 3) and \
                             ('c' in sf_components[2] or
                              'r' in sf_components[2])
        while sf_named_correctly is False:
            subfolder = input(f"""
                            {subfolder} is named incorrectly. 
                            The correct format is:
                            SlabID_GraniteType_NumberOfColumnsOrRows
                            where SlabID_GraniteType_NumberOfColumnsOrRows is of the form:
                            "c#" for columns, or "r#" for rows.
                            Input the correct name: 
                            """)
            sf_components = subfolder.split("_")
            sf_named_correctly = (len(sf_components) == 3) and \
                                 ('c' in sf_components[2] or
                                  'r' in sf_components[2])

    def run(self):
        for parent_folder in self.parent_folders:
            # Printing status
            if parent_folder.index(self.parent_folders) > 1:
                print("")
            print(f"\nEntering into {parent_folder}: parent folder {self.parent_folders.index(parent_folder) + 1} "
                  f"of {len(self.parent_folders)}.")
            # Ensuring that the parent folder was named correctly
            self.parent_folder_named_correctly(parent_folder)
            # Determining image dimensions, crop-necessity from the parent folder name
            self.image_dimensions, self.image_dimensions_mm, self.should_crop_into_four \
                = self.configure_image_dimensions(parent_folder)
            # Getting a list of subfolders in the parent folder
            subfolders = [f.name for f in os.scandir(os.path.join(self.main_folder, parent_folder)) if f.is_dir()]
            # Iterating through subfolders
            for subfolder in subfolders:
                # Printing status
                if subfolders.index(subfolder) > 1:
                    print("")
                print(
                    f"\nEntering into {subfolder}: subfolder {subfolders.index(subfolder) + 1} of {len(subfolders)}.")
                # Ensuring that the subfolder was named correctly
                self.subfolder_named_correctly(subfolder)
                # Getting the path of the current subfolder
                current_folder_path = os.path.join(self.main_folder, parent_folder, subfolder)
                # Making the requisite folders
                exp_folder_path, cropped_folder_path, sim_folder_path, neg_folder_path, pos_folder_path = \
                    self.make_folders(current_folder_path)
                # TODO: TEMPORARY(?)
                # If specified, cropping images, saving to the cropped images folder, and reassigning relevant variable
                if self.should_crop_into_four == 'y':
                    self.crop_into_four(self.get_file_names(current_folder_path), current_folder_path,
                                        cropped_folder_path)
                    current_folder_path = cropped_folder_path
                # Creating experiment images
                exp_csv_file_path = self.create_experiment_images(parent_folder, subfolder, current_folder_path,
                                                                  exp_folder_path)
                # Creating simulation images
                sim_csv_file_path = self.create_simulation_images(exp_folder_path, sim_folder_path)
                # Creating negative images
                neg_csv_file_path = self.create_negative_images(exp_folder_path, neg_folder_path, pos_folder_path)
                # If specified, uploading images
                if self.upload_now == 'y':
                    self.upload_subjects(exp_csv_file_path, sim_csv_file_path, neg_csv_file_path)
        # Pushing updated manifests to Github
        self.push_manifests()
