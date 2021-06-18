import json
import csv
import os
import numpy as np
from shutil import copyfile, rmtree
from PIL import Image
import openpyxl

# from panoptes_client import Panoptes, Project, SubjectSet, Workflow

"""
    (kSWAP instance).users[(user ID)] is an instance of the 'User' class, having attributes:
        user_id, classes, k = len(classes), gamma, user_default, user_score, confusion_matrix, history
        history = [(subject ID), ('user_score')]
            Example: [201, {"0": [0.6, 0.4], "1": [0.3, 0.7]}]
    (kSWAP instance).subjects[(subject ID)] is an instance of the 'Subject' class, having attributes:
        subject_id, score, classes, gold_label, epsilon, retired (boolean), retired_as, seen, history
        history = [(classification ID), (user ID), ('user_score'), (submitted classification), (subject score)]
            Example: [1001, 101, {"0": [0.6, 0.4], "1": [0.3, 0.7]}, 1, {"0": 0.35, "1": 0.65}]
"""


class PromoteSubjects:
    beta = True  # TODO: DELETE
    marking_folder_path = "./markings"
    marking_excel_path = "./manifests/Marking_Manifest.xlsx"
    promotion_threshold_coefficient = 1.5

    def __init__(self, swap, swap_config, classification_csv_path):
        self.swap = swap
        self.swap_config = swap_config
        self.classification_csv_path = classification_csv_path
        self.positive_prior = self.swap_config.p0['1']
        self.promotion_threshold = self.promotion_threshold_coefficient * self.positive_prior
        # """
        # eligible_subjects is a list of dictionaries, one dictionary for each subject whose positive
        # probability is grater than the threshold for promotion to the inspection workflow.
        #     eligible_subjects[i] = {'subject_id': (int) , 'marking_classification_ids': (list of ints) }
        # """
        # self.eligible_subjects = eligible_subjects
        # self.classification_csv_path = classification_csv_path

    @staticmethod
    def parametric_ellipse(cx, cy, xdim, ydim, angle):
        t = np.linspace(0, 2 * np.pi, 1000)
        boundary_x = cx + xdim * np.cos(t) * np.cos(angle) - ydim * np.sin(t) * np.sin(angle)
        boundary_y = cy + xdim * np.cos(t) * np.sin(angle) + ydim * np.sin(t) * np.cos(angle)
        return boundary_x, boundary_y

    @staticmethod
    def ellipse_eq_lhs(x, y, cx, cy, xdim, ydim, angle):
        return (((x - cx) * np.cos(angle) + (y - cy) * np.sin(angle)) / xdim) ** 2 + \
               (((x - cx) * np.sin(angle) - (y - cy) * np.cos(angle)) / ydim) ** 2

    @staticmethod
    def get_interior_pixels(cx, cy, xdim, ydim, angle):
        boundary_x, boundary_y = parametric_ellipse(cx, cy, xdim, ydim, angle)
        pixels = []
        for i in range(int(np.min(boundary_x)), int(np.max(boundary_x)) + 1):
            for j in range(int(np.min(boundary_y)), int(np.max(boundary_y)) + 2):
                pixels.append((i, j))
        interior_pixels = []
        for p in pixels:
            if ellipse_eq_lhs(p[0], p[1], cx, cy, xdim, ydim, angle) <= 1:
                interior_pixels.append(p)
        return interior_pixels

    @staticmethod
    def parse_annotation(annotation):
        value = annotation[0]['value'][0]
        # Center coordinates
        cx = value['x']
        cy = value['y']
        # Axes lengths
        xdim = value['rx']
        ydim = value['ry']
        # Angle wrt +x axis
        angle = value['angle']
        return cx, cy, xdim, ydim, angle

    @staticmethod
    def get_classification_metadata(classification_id, classification_csv_path):
        with open(classification_csv_path, 'r') as read_csv:
            # Create csv reader instance
            csv_reader = csv.DictReader(read_csv)
            # Find, return the specified line of the classifications csv
            for row in csv_reader:
                if int(row['classification_id']) == classification_id:
                    annotations = json.loads(row['annotations'])
                    subject_data = json.loads(row['subject_data'])
        return annotations, subject_data

    @staticmethod
    def get_marking_ids(subjects):
        """Returns a sorted (ascending) list of the classification IDs where a marking was made.
            subjects: list of kSWAP subject instances
        """
        marking_classification_ids = dict((subject, []) for subject in subjects)
        for subject in subjects:
            subject_history = subject.history
            classifications = np.array([sh[-2] for sh in subject_history][1:])
            marking_indices = list(np.where(classifications == 1)[0])
            for mi in marking_indices:
                classification_id = subject_history[mi][0]
                if classification_id != '_':
                    marking_classification_ids[subject].append(classification_id)
        # marking_classification_ids.sort()
        return marking_classification_ids

    @staticmethod
    def get_eligible_subjects(subjects, promotion_threshold):
        """
        Returns a list of subjects whose positive probabilities surpass the promotion threshold.
            subjects: list of kSWAP subject instances
            promotion_threshold: positive probability required for promotion
        """
        for subject in subjects:
            # Get subject's gold labels (-1: non-training, 0: negative, 1: positive)
            gold_label = subject.gold_label
            # Get subject's probability of being 'positive' (of containing a melt-patch)
            positive_probability = json.loads(subject.score)['1']
            # Ignore training subjects
            if gold_label in [0, 1]:
                continue
            if positive_probability > promotion_threshold:
                eligible_subjects.append(subject)
        return eligible_subjects

    @staticmethod
    def get_similar_markings(markings):
        """Markings are assumed to be have been made on the same subject"""
        similar_marking_pairs = []
        for outside in markings:
            for inside in markings:
                if inside['center_coordinates'] in outside['interior_pixels']:
                    similar_marking_pairs.append({outside, inside})
        smp_duplicates_removed = []
        [smp_duplicates_removed.append(x) for x in similar_marking_pairs if x not in smp_duplicates_removed]
        similar_markings = dict((marking, []) for marking in markings)
        for marking_pair in smp_duplicates_removed:
            marking_pair = list(marking_pair)
            marking_pair.sort()
            similar_markings[str(marking_pair[0])].append(marking_pair[1])
        return similar_markings

    @staticmethod
    def get_subject_history(subject, classification_ids):
        subject_history = subject.history
        wanted_subject_history = []
        # Ensuring that histories are stored in the correct order (understood to be inefficient)
        for cl_id in classification_ids:
            for sh in subject_history:
                if int(sh[0]) == int(cl_id):
                    wanted_subject_history.append(sh)
        return wanted_subject_history

    @staticmethod
    def positive_probability_update(prior_positive_probability, user_score):
        """
        Assumes that a marking was made, and updates the subject's probability according
        to the 'score' (confusion matrix) of the user who made the classification.
        """
        user_tpr = user_score["1"][1]
        user_fnr = user_score["1"][0]
        prior = prior_positive_probability
        return (user_tpr * prior) / (user_tpr * prior + user_fnr * (1 - prior))

    def calculate_positive_probability(self, subject_history, prior_positive_probability, ignore_poor=False):
        user_scores = [sh[2] for sh in subject_history]
        positive_probability = [prior_positive_probability]
        i = 0
        for user_score in user_scores:
            positive_probability.append(self.positive_probability_update(positive_probability[i], user_score))
            i += 1
        if ignore_poor is True:
            # Don't let a poor user ruin the subject's probability (this is intentionally opposed to the ethos of SWAP)
            probability_to_return = positive_probability[0]
            for i in range(len(positive_probability) - 1):
                if positive_probability[i + 1] > positive_probability[i]:
                    probability_to_return = positive_probability[i + 1]
        else:
            probability_to_return = positive_probability[-1]
        return probability_to_return

    def parse_subject_data(self, subject_data):
        subject_znv_id = list(subject_data[i].keys())[0]
        subject_dict = subject_data[subject_znv_id]
        subject_id = ubject_dict['!subject_id']
        file_name = subject_dict['#file_name']
        if self.beta is True:
            subfolder_name = 'beta_experiment'
            parent_folder_name = 'beta_images'
        else:
            try:
                subfolder_name = subject_dict['#subfolder_name']
                parent_folder_name = subject_dict['#parent_folder_name']
            except KeyError as e:
                print(f'Error {e} with folder names.')
                subfolder_name = None
                parent_folder_name = None
        warehouse = subject_dict['#warehouse']
        location = subject_dict['#location']
        latitude_longitude = subject_dict['#latitude_longitude']
        slab_id = subject_dict['#slab_id']
        number_of_columns = subject_dict['#number_of_columns']
        return subject_id, file_name, subfolder_name, parent_folder_name, warehouse, \
               location, latitude_longitude, slab_id, number_of_columns

    def get_markings_to_promote(self, marking_classification_ids):
        subjects = list(marking_classification_ids.keys())
        for subject in subjects:
            for mc_id in marking_classification_ids[subject]:
                annotation, subject_data = self.get_classification_metadata(mc_id, self.classification_csv_path)
                cx, cy, xdim, ydim, angle = self.parse_annotation(annotation)
                interior_pixels = self.get_interior_pixels(cx, cy, xdim, ydim, angle)
                markings.append({'subject': subject, 'classification_id': mc_id, 'subject_data': subject_data,
                                 'cx': cx, 'cy': cy, 'xdim': xdim, 'ydim': ydim, 'angle': angle,
                                 'interior_pixels': interior_pixels})
            similar_markings = self.get_similar_markings(markings)
            eligible_markings = similar_markings.keys()
            markings_to_promote = []
            for marking in eligible_markings:
                if similar_markings[marking]:
                    classification_ids = [marking['classification_id'], *[sm['classification_id'] for
                                                                          sm in similar_markings[marking]]]
                    subject_marking_history = self.get_subject_history(subject, classification_ids)
                    positive_probability = self.calculate_positive_probability(subject_marking_history,
                                                                               self.positive_prior,
                                                                               ignore_poor=True)
                    if positive_probability > self.promotion_threshold:
                        marking_pool = [marking, *eligible_markings[marking]]
                        cxs = [m['cx'] for m in marking_pool]
                        cys = [m['cy'] for m in marking_pool]
                        xdims = [m['xdim'] for m in marking_pool]
                        ydims = [m['ydim'] for m in marking_pool]
                        angles = [m['angle'] for m in marking_pool]
                        average_x = sum(cxs) / len(cxs)
                        average_y = sum(cys) / len(cys)
                        average_xdim = sum(xdims) / len(xdims)
                        average_ydim = sum(ydims) / len(ydims)
                        average_angle = sum(angles) / len(angles)
                        for i in range(len(marking_pool) - 1):
                            assert (marking_pool[i]['subject_data'] == marking_pool[i + 1]['subject_data'])
                        subject_data = marking_pool[0]['subject_data']
                        original_subject_id, original_file_name, subfolder_name, parent_folder_name, warehouse, \
                        location, latitude_longitude, slab_id, number_of_columns = self.parse_subject_data(subject_data)
                        markings_to_promote.append({'original_subject_id': original_subject_id,
                                                    'original_file_name': original_file_name,
                                                    'subfolder_name': subfolder_name,
                                                    'parent_folder_name': parent_folder_name,
                                                    'warehouse': warehouse,
                                                    'location': location,
                                                    'latitude_longitude': latitude_longitude,
                                                    'slab_id': slab_id,
                                                    'number_of_columns': number_of_columns,
                                                    'average_x': average_x,
                                                    'average_y': average_y,
                                                    'average_xdim': average_xdim,
                                                    'average_ydim': average_ydim,
                                                    'average_angle': average_angle,
                                                    'number': len(marking_pool)})
            return markings_to_promote

    @staticmethod
    def configure_csv(folder_path, desired_csv_file_name, metadata_fields_list=None):
        if metadata_fields_list is None:
            metadata_fields_list = ['']
        desired_csv_file_path = os.path.join(folder_path, desired_csv_file_name)
        with open(desired_csv_file_path, 'w', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields_list)
            csv_writer.writeheader()
        return desired_csv_file_path

    @staticmethod
    def configure_excel(excel_file_path):
        wb = openpyxl.load_workbook(filename=excel_file_path)
        ws = wb['Sheet1']
        first_empty_row = None
        for row in range(1, int(1e10)):
            if ws.cell(row, 1).value is None:
                first_empty_row = row
                break
        return wb, ws, first_empty_row

    @staticmethod
    def configure_marking_metadata(markings, mrk_i):
        """markings is a list of dicts in the form of markings_to_promote"""
        marking_metadata = []
        for marking in markings:
            subject_id = "m" + str(int(mrk_i) + markings.index(marking))
            file_name = marking_id + "_" + marking['original_file_name']
            marking_metadata.append(
                [subject_id, file_name, marking["subfolder_name"], marking["parent_folder_name"],
                 marking["original_subject_id"], marking["original_file_name"], marking["warehouse"], marking["location"],
                 marking["latitude_longitude"], marking["slab_id"], marking["number"], marking["average_x"],
                 marking["average_y"], marking["average_xdim"], marking["average_ydim"], marking["average_angle"]])
        return marking_metadata

    @staticmethod
    def write_metadata_into_csv(csv_file_path, metadata_fields, metadata_list):
        with open(csv_file_path, 'a', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)
            for j in range(len(metadata_list)):
                row = {}
                for i in range(len(metadata_fields)):
                    row[str(metadata_fields[i])] = metadata_list[i]
                csv_writer.writerow(row)

    @staticmethod
    def write_metadata_into_manifest(ws, starting_row, metadata_list):
        row = starting_row
        for j in range(len(metadata_list)):
            for i in range(len(metadata_list[j])):
                ws.cell(row=row, column=(i+1)).value = str(metadata_list[j][i])
            row += 1

    def fetch_images_from_google_drive(self):
        pass

    @staticmethod
    def get_file_names(folder_path):
        all_file_names = os.listdir(folder_path)
        wanted_file_names = []
        for file_name in all_file_names:
            if file_name.endswith(".jpeg") or file_name.endswith(".jpg") or file_name.endswith(".png"):
                wanted_file_names.append(file_name)
        return wanted_file_names

    @staticmethod
    def create_marking_images(markings, marked_images):
        """markings is a list of dicts in the form of markings_to_promote"""
        pass

    def run(self):
        # Get list of 'Subject' instances (each one corresponding to a Zooniverse subject)
        subjects = list(self.swap.subjects.values())
        # Get a list of subjects whose positive probability surpass the probability threshold
        eligible_subjects = self.get_eligible_subjects(subjects, self.promotion_threshold)
        # Get a dict of the classification IDs where a marking was made for each eligible subject
        markings_classification_ids = self.get_marking_ids(eligible_subjects)
        # Get a dict of markings that themselves pass the probability threshold
        markings_to_promote = self.get_markings_to_promote(markings_classification_ids)
        # Make a folder for promoted markings
        os.makedirs(self.marking_folder_path, exist_ok=True)
        # Configure markings' excel manifest and .csv file
        marking_metadata_fields = ["!subject_id", "#file_name", "#subfolder_name", "#parent_folder_name",
                                   "#original_subject_id", "original_file_name", "#warehouse", "#location",
                                   "latitude_longitude", "slab_id", "#number_of_columns", "#number_of_times_marked",
                                   "average_x", "average_y", "average_xdim", "average_ydim", "average_angle"]
        wb, ws, mrk_i = self.configure_excel(self.marking_excel_path)
        marking_csv_path = self.configure_csv(self.marking_folder_path, "marking_subjects.csv", marking_metadata_fields)
        marking_metadata = self.configure_marking_metadata(markings_to_promote, mrk_i)
        # Write markings' metadata into excel manifest and .csv file
        self.write_metadata_into_csv(marking_csv_path, marking_metadata_fields, marking_metadata)
        self.write_metadata_into_manifest(ws, mrk_i, marking_metadata)
        # CONTINUE HERE
        # Fetch images from Google Drive and copy them into the markings folder
        self.fetch_images_from_google_drive()
        # Get files from the markings folder
        marked_images = self.get_file_names(self.marking_folder_path)
        # Crop images around the markings, draw average markings, save to the marking folder
        self.create_marking_images(markings_to_promote, marked_images)
        # Promote the marking obtained above to the inspection workflow
        """
        upload images & .csv to zooniverse
        upload markings to google drive, delete from local
        upload manifests to github OR to google drive
        """
