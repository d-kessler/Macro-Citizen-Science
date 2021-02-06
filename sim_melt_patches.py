# importing cv2 
import cv2
import cv
import argparse
import random
import numpy as np
from matplotlib import pyplot as plt
from skimage.color import rgb2lab, deltaE_cie76

import sys
import os
import io
import csv
import openpyxl
from PIL import Image, ExifTags

from config import *
from image_tools import *


def configure_file_paths():
    """Get image folder path, create folder for simulations"""

    image_folder_path = input('Enter the folder path: ')
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Data, Analysis\Slab 3 6x8 no flash"
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Data, Analysis\Slab 1 6x8 no flash"

    return image_folder_path


def make_sim_folder(image_folder_path_):
    """Create a folder for simulation images"""
    sim_folder_path = os.path.join(image_folder_path_, r"simulations")
    try:
        os.mkdir(sim_folder_path)
    except FileExistsError:
        clear_folder(sim_folder_path)

    return sim_folder_path


def configure_metadata():
    feedback_id = 'meltpatch'
    training_subject = 'true'

    return feedback_id, training_subject


def configure_csv(sim_folder_path_):
    """Configuring csv file, saving in simulation images folder"""
    csv_file_name = 'simulation_subjects.csv'
    csv_file_path = os.path.join(sim_folder_path_, csv_file_name)

    with open(csv_file_path, 'w', newline='') as f:
        metadata_fields = ['!subject_id', '#file_name', '#training_subject', '#feedback_1_id', '#feedback_1_x',
                           '#feedback_1_y', '#feedback_1_toleranceA', '#feedback_1_toleranceB', '#feedback_1_theta',
                           '#minor_to_major_ratio']
        csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)
        csv_writer.writeheader()

    return csv_file_name, csv_file_path, metadata_fields


def write_metadata_into_excel(ws, i, subject_id, sim_file_name, training_subject, feedback_id, center_coordinates,
                              axesLength,
                              angle, minor_to_major_ratio):
    ws.cell(row=i, column=1).value = subject_id
    ws.cell(row=i, column=2).value = str(sim_file_name)
    ws.cell(row=i, column=3).value = training_subject
    ws.cell(row=i, column=4).value = feedback_id
    ws.cell(row=i, column=5).value = center_coordinates[0]
    ws.cell(row=i, column=6).value = center_coordinates[1]
    ws.cell(row=i, column=7).value = axesLength[0]
    ws.cell(row=i, column=8).value = axesLength[1]
    ws.cell(row=i, column=9).value = angle
    ws.cell(row=i, column=10).value = minor_to_major_ratio


def write_metadata_into_csv(csv_file_path_, metadata_fields_, subject_id, sim_file_name, training_subject, feedback_id,
                            center_coordinates, axesLength,
                            angle, minor_to_major_ratio):
    with open(csv_file_path_, 'a', newline='') as f:
        csv_writer = csv.DictWriter(f, fieldnames=metadata_fields_)

        row = {'!subject_id': subject_id}
        row['#file_name'] = str(sim_file_name)
        row['#training_subject'] = training_subject
        row['#feedback_1_id'] = feedback_id
        row['#feedback_1_x'] = center_coordinates[0]
        row['#feedback_1_y'] = center_coordinates[1]
        row['#feedback_1_toleranceA'] = axesLength[0]
        row['#feedback_1_toleranceB'] = axesLength[1]
        row['#feedback_1_theta'] = angle
        row['#minor_to_major_ratio'] = minor_to_major_ratio
        csv_writer.writerow(row)


def create_sims_from_process_images(folder_path, upload_now_, lower_limit_, max_sample_, simulation_set_id_=None):
    image_folder_path = folder_path
    lower_limit = lower_limit_
    upload_now = upload_now_
    simulation_set_id = simulation_set_id_
    max_sample = max_sample_

    sim_file_paths = draw_sims(image_folder_path, lower_limit, max_sample, upload_now,
                               simulation_set_id_=simulation_set_id)

    # Ensuring that scale bars weren't drawn over
    for sim_file_path in sim_file_paths:
        image = Image.open(sim_file_path)
        image_exif = image.getexif()

        draw_scale_bars(sim_file_path, image, image_exif, lower_limit)

    return sim_file_paths


def draw_sims(image_folder_path, lower_limit_, max_sample, upload_now_, simulation_set_id_=None):
    # Create a folder for simulated images within the folder where images are fetched
    sim_folder_path = make_sim_folder(image_folder_path)  # (specified folder with images)

    # Configure excel file to be written into, find the first empty row
    excel_file_path = r"manifests/Simulation_Manifest.xlsx"
    wb, ws, first_empty_row = configure_excel(excel_file_path)

    # Fetch predefined metadata fields
    feedback_id, training_subject = configure_metadata()

    # Create csv file; get its name, file path, and metadata fields (column headers)
    csv_file_name, csv_file_path, metadata_fields = configure_csv(sim_folder_path)

    # Get list of file names in a directory
    file_names = get_file_names(image_folder_path)

    # Sample from the file names list at most 'max_sample' # of files, store in new list
    if len(file_names) >= max_sample:
        sample_number = max_sample
    else:
        sample_number = len(file_names)
    select_file_names = sample_from_file_names(file_names, sample_number)  # (list to sample from, number to sample)

    # Index variable used to assign subject ID, decide value of axesLength
    sim_file_paths = []
    # TODO: uncomment
    # i = first_empty_row
    i = 1

    # Iterating through sampled files
    for filename in select_file_names:

        # Assigning a subject ID equal to 's' (for simulation) plus the total number of such subjects as of this one's addition
        subject_id = 's' + str(i - 1)

        # Getting the file's full path, creating a path for the simulated image to be created from it
        image_file_path = os.path.join(image_folder_path, filename)
        sim_file_name = subject_id + r"_" + filename
        sim_file_path = os.path.join(sim_folder_path, sim_file_name)

        sim_file_paths.append(sim_file_path)

        # Creating an cv2 image instance named 'image'
        image = cv2.imread(image_file_path)

        # Getting image height and width (in units of pixels)
        height, width, _ = image.shape

        # Find the darkest color present in the image
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        minVal = cv2.minMaxLoc(gray_image)[0]

        # Finding the number of millimeters (on the image's 6x8 inch scale) per pixel
        mm_per_pixel = get_mm_per_pixel(image)  # (image variable)

        # Setting ellipse center coordinates and angle w.r.t. positive x-axis
        center_x = random.randint(int(25/mm_per_pixel), int((width-25)/mm_per_pixel))
        center_y = random.randint(int(25/mm_per_pixel), int((height-25)/mm_per_pixel))
        center_coordinates = (center_x, center_y)
        angle = random.randint(0, 360)
        startAngle = 0
        endAngle = 360

        # # SWAP Continuum choice of minor axes lengths
        # minor_axes_distribution = np.concatenate(
        #     (np.linspace(1, 1.5, 20), np.linspace(1.5, 2, 20), np.linspace(2, 3, 20),
        #      np.linspace(3, 4, 20), np.linspace(4, 6, 20)))
        # minor_axis_mm = random.choice(minor_axes_distribution)
        # minor_axis = minor_axis_mm / mm_per_pixel

        # SWAP Discrete choice of axes lengths
        #
        # minor_axes_mm = [2, 3]  # minor axes lengths in millimeters
        # minor_axes = [minor_axis / mm_per_pixel for minor_axis in minor_axes_mm]
        #
        # # Creating an equal numbers of ellipses of lower limit and 3mm semi-minor axes lengths
        # if i % 2 == 0:
        #     axesLength = [int(major_axes[0]), int(minor_axes[0])]
        # else:
        #     axesLength = [int(major_axes[1]), int(minor_axes[1])]

        # # Setting major axis
        # minor_to_major_ratio = random.uniform(1, 10)
        # if minor_to_major_ratio > 10:
        #     minor_to_major_ratio = 10
        # major_axis = minor_axis * minor_to_major_ratio
        # axesLength = [major_axis, minor_axis]

        if i in range(1, 10) or i == 1 or i == 10:
            minor_axis = 1 / mm_per_pixel
            minor_to_major_ratio = i

        if i in range(11, 20) or i == 11 or i == 20:
            minor_axis = 1.5 / mm_per_pixel
            minor_to_major_ratio = i - 10

        if i in range(21, 30) or i == 21 or i == 30:
            minor_axis = 2 / mm_per_pixel
            minor_to_major_ratio = i - 20

        if i in range(31, 40) or i == 31 or i == 40:
            minor_axis = 2.5 / mm_per_pixel
            minor_to_major_ratio = i - 30

        if i in range(41, 50) or i == 41 or i == 50:
            minor_axis = 3 / mm_per_pixel
            minor_to_major_ratio = i - 40

        if i in range(51, 60) or i == 51 or i == 60:
            minor_axis = 3.5 / mm_per_pixel
            minor_to_major_ratio = i - 50

        if i in range(61, 65) or i == 61 or i == 65:
            minor_axis = 4 / mm_per_pixel
            minor_to_major_ratio = i - 60

        if i in range(66, 70) or i == 66 or i == 70:
            minor_axis = 4.5 / mm_per_pixel
            minor_to_major_ratio = i - 65

        if i in range(71, 73) or i == 71 or i == 73:
            minor_axis = 5 / mm_per_pixel
            minor_to_major_ratio = i - 70

        if i in range(74, 77) or i == 74 or i == 77:
            minor_axis = 5.5 / mm_per_pixel
            minor_to_major_ratio = i - 73

        if i in range(78, 80) or i == 78 or i == 80:
            minor_axis = 6 / mm_per_pixel
            minor_to_major_ratio = i - 77

        major_axis = minor_axis * minor_to_major_ratio
        name = '{:.1f}mm_{:.1f}mm_{:.1f}'.format(minor_axis * mm_per_pixel, int(major_axis * mm_per_pixel),
                                                 minor_to_major_ratio)

        axesLength = [major_axis, minor_axis]

        # Getting axes radii (the input for cv2 ellipse function)
        axesRadii = tuple([int(j / 2) for j in axesLength])

        # Setting the color of the drawn ellipse equal to the dark color in the image
        color = (minVal, minVal, minVal)

        # Draw (filled-in) ellipse with specified parameters
        image = cv2.ellipse(image, center_coordinates, axesRadii, angle,
                            startAngle, endAngle, color, thickness=-1, lineType=cv2.LINE_AA)
        image = cv2.circle(image, center_coordinates, 500, (0, 255, 0), 10)
        # image = cv2.putText(image, '{:.2f}mm, {:.2f}, {:.2f}'.format(np.ceil(2 * axesRadii[1] * mm_per_pixel), int(2 * axesRadii[0] * mm_per_pixel), minor_to_major_ratio),
        #                     (center_coordinates[0] - 120, center_coordinates[1] - 120), cv2.FONT_HERSHEY_SIMPLEX, 2,
        #                     (0, 255, 0), thickness=4)

        # Saving simulated image to its created filed path
        # TODO: uncomment
        # cv2.imwrite(sim_file_path, image)
        save_name = os.path.join(sim_folder_path, name + r".jpeg")
        cv2.imwrite(save_name, image)

        # # Draw a circle around the drawn ellipse in green (for checking purposes)
        # sim_file_path_circled = os.path.join(sim_folder_path, subject_id + r"_CircledSim.jpeg")
        # circled_image = cv2.circle(image, center_coordinates, 4 * axesLength[0], (0, 255, 0), 10)
        # cv2.imwrite(sim_file_path_circled, circled_image)

        # TODO: uncomment, change last argument
        # image = Image.open(sim_file_path)
        # image_exif = image.getexif()
        # draw_scale_bars(sim_file_path, image, image_exif, 2)

        # Resizing the image to be under the Zooniverse recommended 600KB limit
        # TODO: uncomment
        # resize_to_limit(sim_file_path)

        # Write metadata values into both the specified excel file and created csv
        # write_metadata_into_excel(ws, i, subject_id, sim_file_name, training_subject, feedback_id, center_coordinates,
        #                           axesLength, angle, minor_to_major_ratio)
        write_metadata_into_csv(csv_file_path, metadata_fields, subject_id, sim_file_name, training_subject,
                                feedback_id, center_coordinates,
                                axesLength, angle, minor_to_major_ratio)

        # TODO: uncomment
        print('\r{} of {} simulations made.'.format((select_file_names.index(filename) + 1), len(select_file_names)),
              end="")
        i += 1

    # Saving the excel manifest --- it must be closed
    wb.save(excel_file_path)

    # Executing terminal commands to upload images
    if upload_now_ == 'y':
        cmd = 'panoptes subject-set upload-subjects {} {}'.format(simulation_set_id_, csv_file_path)
        os.system(cmd)
        print('\nSimulation subjects uploaded.')

    return sim_file_paths


def run():
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Python\Macro-Citizen-Science\images\beta-sample"
    image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Python\Macro-Citizen-Science\images\beta-sample-2"
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Python\Macro-Citizen-Science\images\Slab 1 6x8 no flash - Copy"
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Python\Macro-Citizen-Science\images\1Slab_1_12x16_no_flash\cropped"
    # image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Python\Macro-Citizen-Science\images\slab 3"
    lower_limit = []
    max_sample = 100
    upload_now = 'n'

    draw_sims(image_folder_path, lower_limit, max_sample, upload_now)


run()
