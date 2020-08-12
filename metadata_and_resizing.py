import sys
import os
import io
import json
import csv
import openpyxl
from PIL import Image, ExifTags
from pathlib import Path
from datetime import datetime
from panoptes_client import Panoptes, Project, SubjectSet, Subject
import image_slicer


def configure():
    """Get file & slab information, create new folder for images, configure excel & csv files"""
    global subject_set_id, image_folder_path, excel_file_path, warehouse_name, location, granite_type, slab_id, should_crop_into_four,\
        cropped_folder_path, resized_folder_path, csv_file_name, csv_file_path, metadata_fields, wb, ws, first_empty_row

    subject_set_id = 86450

    # # getting user input for metadata fields
    # image_folder_path = input('Enter the folder path: ')
    # excel_file_path = input('Excel file path: ')
    # warehouse_name = input('Warehouse name: ')
    # location = input('Location (City, State): ')
    # granite_type = input('Granite type: ')
    # slab_id = input('Slab ID: ')
    should_crop_into_four = input('Should the images be cropped into 4 parts? Enter \'yes\' or \'no\' :  ')

    image_folder_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Data, Analysis\Slab 1 12x16 no flash - Copy"
    excel_file_path = r"C:\Users\dkess\OneDrive\Documents\CWRU\Macro\Data, Analysis\Experiment_Manifest.xlsx"
    warehouse_name = 'United Stone International'
    location = 'Solon, Ohio'
    granite_type = 'Dallas White'
    slab_id = '1151|20'

    # creating a folder for cropped images
    cropped_folder_path = os.path.join(image_folder_path, r"cropped_" + os.path.basename(image_folder_path))
    try:
        os.mkdir(cropped_folder_path)
    except:
        print('Cropped Folder Exists.')

    # creating a folder for resized images
    resized_folder_path = os.path.join(image_folder_path, r"resized_" + os.path.basename(image_folder_path))
    try:
        os.mkdir(resized_folder_path)
    except:
        print('Resized Folder Exists.')

    # configuring excel file, finding first empty row
    wb = openpyxl.load_workbook(filename=excel_file_path)
    ws = wb['Sheet1']
    for cell in ws["B"]:
        if cell.value is None:
            first_empty_row = cell.row
            break

    # configuring csv file, saving in resized images folder
    csv_file_name = 'experiment_subjects_csv.csv'
    csv_file_path = os.path.join(resized_folder_path, csv_file_name)

    with open(csv_file_path, 'w', newline='') as f:
        metadata_fields = ['!subject_id', '#file_name', '#warehouse', '#location', '#granite_type', '#slab_id',
                           '#date_time',
                           '#latitude_longitude']
        csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)
        csv_writer.writeheader()


def get_file_names():
    """Create a list of image files in given directory"""
    global file_names

    all_file_names = os.listdir(image_folder_path)
    file_names = []
    for file in all_file_names:
        if file.endswith(".jpeg") or file.endswith(".jpg") or file.endswith(".png"):
            file_names.append(file)


def crop_into_four(pil_file_):
    global cropped_file_paths, resized_file_paths

    cropped_file_paths = []
    resized_file_paths = []

    width, height = pil_file_.size

    half_width = width / 2
    half_height = height / 2

    # starting from top left (0,0) and moving clockwise
    # (left, top, right, bottom)
    section_1 = (0, 0, half_width, half_height)
    section_2 = (half_width, 0, width, half_height)
    section_3 = (half_width, half_height, width, height)
    section_4 = (0, half_height, half_width, height)

    for i in range(1, 5):
        im = pil_file_

        if i == 1:
            im = im.crop(section_1)
        if i == 2:
            im = im.crop(section_2)
        if i == 3:
            im = im.crop(section_3)
        if i == 4:
            im = im.crop(section_4)


        reformatted_cropped_file_path = cropped_file_path.replace(str(extension), "_{}{}".format(str(i), str(extension)))
        reformatted_resized_file_path = resized_file_path.replace(str(extension), "_{}{}".format(str(i), str(extension)))

        cropped_file_paths.append(reformatted_cropped_file_path)
        resized_file_paths.append(reformatted_resized_file_path)

        im = im.save(reformatted_cropped_file_path, exif=image_exif)


def resize_to_limit(image_file_path_=None, resized_file_path_=None, size_limit=600000):
    """Resize images to size_limit, save to new file"""
    # add conditional for should_resize

    if should_crop_into_four == 'yes':

        for cropped_file in cropped_file_paths:
            index = list.index(cropped_file_paths, cropped_file)

            img = Image.open(cropped_file)
            aspect = img.size[0] / img.size[1]

            while True:
                with io.BytesIO() as buffer:
                    img.save(buffer, format="JPEG")
                    data = buffer.getvalue()
                filesize = len(data)
                size_deviation = filesize / size_limit
                print("size: {}; factor: {:.3f}".format(filesize, size_deviation))

                if size_deviation <= 1:
                    img.save(resized_file_paths[index], exif=image_exif)
                    print(cropped_file + '\nsaved to\n' + resized_file_paths[index])
                    # print(img.size[0], img.size[1])
                    break
                else:
                    new_width = img.size[0] / (1 * (size_deviation ** 0.5))
                    new_height = new_width / aspect

                    img = img.resize((int(new_width), int(new_height)))
    else:
        img = img_orig = Image.open(image_file_path_)
        img_exif = img.info['exif']
        aspect = img.size[0] / img.size[1]

        while True:
            with io.BytesIO() as buffer:
                img.save(buffer, format="JPEG")
                data = buffer.getvalue()
            filesize = len(data)
            size_deviation = filesize / size_limit
            # print("size: {}; factor: {:.3f}".format(filesize, size_deviation))

            if size_deviation <= 1:
                img.save(resized_file_path_, exif=img_exif)
                # print(img.size[0], img.size[1])
                break
            else:
                new_width = img.size[0] / (1 * (size_deviation ** 0.5))
                new_height = new_width / aspect

                img = img_orig.resize((int(new_width), int(new_height)))

def get_date(exif_dict):
    """Get date/time image exif data (if available)"""
    global date, has_date
    date = ''
    has_date = 0

    try:
        date = datetime.strptime(exif_dict['DateTimeOriginal'], '%Y:%m:%d %H:%M:%S')
        date = json.dumps(date, default=str)
        has_date = 1

        return date, has_date

    except:
        pass


def get_gps(exif_dict):
    """Get GPS image exif data (if available)"""
    global latitude, longitude, has_gps
    latitude = []
    longitude = []
    has_gps = 0

    gps_dict = {}
    try:
        for key in exif_dict['GPSInfo'].keys():
            gps_tag = ExifTags.GPSTAGS.get(key)
            gps_dict[gps_tag] = exif_dict['GPSInfo'][key]

        latitude_raw = gps_dict.get('GPSLatitude')
        longitude_raw = gps_dict.get('GPSLongitude')

        lat_ref = gps_dict.get('GPSLatitudeRef')
        long_ref = gps_dict.get('GPSLongitudeRef')

        if lat_ref == "S":
            latitude = -abs(convert_to_degrees(latitude_raw))
        else:
            latitude = convert_to_degrees(latitude_raw)

        if long_ref == "W":
            longitude = -abs(convert_to_degrees(longitude_raw))
        else:
            longitude = convert_to_degrees(longitude_raw)

        has_gps = 1

        return latitude, longitude, has_gps

    except:
        pass


def convert_to_degrees(value):
    """Convert GPS exif data to latitude/longitude degree format"""
    d0 = value[0][0]
    d1 = value[0][1]
    d = float(d0) / float(d1)

    m0 = value[1][0]
    m1 = value[1][1]
    m = float(m0) / float(m1)

    s0 = value[2][0]
    s1 = value[2][1]
    s = float(s0) / float(s1)

    return d + (m / 60.0) + (s / 3600.0)


def write_metadata_into_excel():
    ws.cell(row=i, column=1).value = subject_id
    ws.cell(row=i, column=2).value = str(resized_file_name)
    ws.cell(row=i, column=3).value = warehouse_name
    ws.cell(row=i, column=4).value = location
    ws.cell(row=i, column=5).value = granite_type
    ws.cell(row=i, column=6).value = slab_id
    if has_date == 1:
        ws.cell(row=i, column=7).value = date.replace("\"", "")
    elif has_date == 0:
        ws.cell(row=i, column=7).value = '-'
    if has_gps == 1:
        ws.cell(row=i, column=8).value = str(latitude) + ', ' + str(longitude)
    elif has_gps == 0:
        ws.cell(row=i, column=8).value = '-'


def write_metadata_into_csv():
    with open(csv_file_path, 'a', newline='') as f:
        csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)

        row = {'!subject_id': subject_id}
        row['#file_name'] = str(resized_file_name)
        row['#warehouse'] = warehouse_name
        row['#location'] = location
        row['#granite_type'] = granite_type
        row['#slab_id'] = slab_id
        if has_date == 1:
            row['#date_time'] = date.replace("\"", "")
        elif has_date == 0:
            row['#date_time'] = '-'
        if has_gps == 1:
            row['#latitude_longitude'] = str(latitude) + ', ' + str(longitude)
        elif has_gps == 0:
            row['#latitude_longitude'] = '-'

        csv_writer.writerow(row)


configure()
get_file_names()

# crop (if applicable) and resize images, get and fill metadata into excel file & csv
i = first_empty_row
for filename in file_names:
    subject_id = 'e' + str(i - 1)
    extension = os.path.splitext(filename)[-1]

    resized_file_name = r"res-" + filename

    image_file_path = os.path.join(image_folder_path, filename)
    cropped_file_path = os.path.join(cropped_folder_path, filename)
    resized_file_path = os.path.join(resized_folder_path, resized_file_name)

    # creating PIL instance
    pil_file = Image.open(image_file_path)
    image_exif = pil_file.info['exif']

    if should_crop_into_four == 'yes':
        crop_into_four(pil_file)
        resize_to_limit()
    else:
        resize_to_limit(image_file_path, resized_file_path)

    # getting image exif data
    exif = {ExifTags.TAGS[k]: v for k, v in pil_file._getexif().items() if k in ExifTags.TAGS}

    get_date(exif)
    get_gps(exif)

    write_metadata_into_excel()
    write_metadata_into_csv()

    print('{} of {} completed.'.format((file_names.index(filename) + 1), len(file_names)))

    i += 1

# save excel manifest --- the excel file must be closed
wb.save(excel_file_path)

print(
    '\nTo upload subjects, \nEnter into the command line: \n    chdir {}\n\nFollowed by: \n    panoptes subject-set upload-subjects {} {}'.format(
        resized_folder_path, subject_set_id, csv_file_name))
