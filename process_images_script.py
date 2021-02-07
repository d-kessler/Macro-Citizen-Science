import csv
import io
import json
import os
import random
import shutil
import statistics
from datetime import datetime, date

import cv2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from PIL import Image, ImageDraw, ExifTags
from git import Repo
from imutils import perspective, contours, grab_contours
from panoptes_client import Panoptes, Project, SubjectSet, \
    Workflow
from panoptes_client.panoptes import PanoptesAPIException
from scipy.spatial import distance
from skimage.util import img_as_ubyte


class ProcessImages:
    # Training probabilities in the form [[probability] * number of images for which probability is applied]
    training_chances = [[0.40] * 50, [0.20] * 50]
    training_default = [0.10]

    # Image dimensions (height, width) in inches
    image_dimensions = [6, 8]

    # Conversion to millimeters
    mm_per_inch = 25.4
    image_dimensions_mm = [x * 25.4 for x in image_dimensions]

    # Number of training images (simulations and negatives each) made per folder
    # TODO: CHANGE TO BETA-SAMPLE NUMBER
    training_images_per_folder = 5

    # TODO: TEMPORARY
    lower_limit = 2

    def __init__(self, username, password, project_id, workflow_id, parent_folder="images",
                 exp_manifest_path=r"manifests/Experiment_Manifest.xlsx",
                 sim_manifest_path=r"manifests/Simulation_Manifest.xlsx",
                 neg_manifest_path=r"manifests/Negative_Manifest.xlsx",
                 sim_feedback_id="meltpatch", neg_feedback_id="no_meltpatch",
                 sim_template_path=r"sim_tools\sim_template.png"):

        # Initializing Zooniverse username, password, project & workflow IDs
        self.username = username
        self.password = password
        self.project_id = project_id
        self.workflow_id = workflow_id

        # Setting image folders' parent folder
        self.parent_folder = parent_folder

        # Setting experiment, simulation, and negative manifest paths
        self.exp_manifest_path = exp_manifest_path
        self.sim_manifest_path = sim_manifest_path
        self.neg_manifest_path = neg_manifest_path

        # Configuring workflow-level training image feedback IDs
        self.sim_feedback_id = sim_feedback_id
        self.neg_feedback_id = neg_feedback_id

        # Configuring the simulation template's file path
        self.sim_template_path = sim_template_path

        # Getting subject set info (if subjects are being uploaded now)
        self.upload_now = input('Are you looking to upload these subjects now? [y/n]: ')
        while self.upload_now != 'y' and self.upload_now != 'n':
            print('Please enter \'y\' or \'no\'... ')
            self.upload_now = input('Are you looking to upload these subjects now? [y/n]: ')
        if self.upload_now == 'y':
            self.subject_set_ids = self.get_existing_subject_sets()
            self.experiment_set_id, self.need_new_exp = self.configure_subject_set('experiment', self.subject_set_ids)
            self.simulation_set_id, self.need_new_sim = self.configure_subject_set('simulation', self.subject_set_ids)
            self.negative_set_id, self.need_new_neg = self.configure_subject_set('negative', self.subject_set_ids)

            # Configuring designator for selected/created subject sets (possibly redundant, but not time consuming)
            self.configure_designator()

        else:
            self.experiment_set_id = None
            self.simulation_set_id = None
            self.negative_set_id = None

        # Getting a list of subfolders in the parent folder
        self.subfolders = [f.name for f in os.scandir(self.parent_folder) if f.is_dir()]

        # Iterating through subfolders
        for subfolder in self.subfolders:
            print(
                f"\nEntering into {subfolder}: folder {self.subfolders.index(subfolder) + 1} of {len(self.subfolders)}.")

            # Getting the path of the current subfolder
            self.current_folder_path = os.path.join(self.parent_folder, subfolder)

            # Making the necessary folders
            self.should_crop_into_four = input('Should the images be cropped into 4 parts? [y/n]: ')
            while self.should_crop_into_four != 'y' and self.should_crop_into_four != 'n':
                print('Please enter \'y\' or \'no\'... ')
                self.should_crop_into_four = input('Should the images be cropped into 4 parts? [y/n]: ')

            self.exp_folder_path, self.cropped_folder_path, self.sim_folder_path, self.neg_folder_path, self.pos_folder_path \
                = self.make_folders()

            # TODO: TEMPORARY
            self.should_draw_scale_bars = input('Should scale bars be drawn on the images? [y/n]: ')
            while self.should_draw_scale_bars != 'y' and self.should_draw_scale_bars != 'n':
                print('Please enter \'y\' or \'no\'... ')
                self.should_draw_scale_bars = input('Should scale bars be drawn on the images? [y/n]: ')

            # Configuring manifests
            self.exp_wb, self.exp_ws, self.exp_first_empty_row = self.configure_excel(self.exp_manifest_path)
            self.sim_wb, self.sim_ws, self.sim_first_empty_row = self.configure_excel(self.sim_manifest_path)
            self.neg_wb, self.neg_ws, self.neg_first_empty_row = self.configure_excel(self.neg_manifest_path)

            # Configuring experiment images' metadata
            self.warehouse_name, self.location, self.granite_type, self.slab_id, self.number_of_columns = self.configure_exp_metadata()

            # Configuring .CSV files (used for syncing images with metadata in Zooniverse databases)
            self.exp_csv_file_name, self.exp_csv_file_path, self.exp_metadata_fields, self.sim_csv_file_name, \
                self.sim_csv_file_path, self.sim_metadata_fields, self.neg_csv_file_name, self.neg_csv_file_path, \
                self.neg_metadata_fields = self.configure_CSVs()

            # Getting a list of file names in the current folder
            self.image_file_names = self.get_file_names(self.current_folder_path)

            # If specified, cropping images, saving to the cropped images folder, and reassigning relevant variables
            if self.should_crop_into_four == 'y':
                self.current_folder_path, self.image_file_names = self.crop_into_four()

            # Initializing indexing variables and other variables used to position images on the slab
            # self.i = self.exp_first_empty_row
            self.exp_i = self.exp_first_empty_row
            self.sim_i = self.sim_first_empty_row
            self.neg_i = self.neg_first_empty_row
            self.row = 1
            self.col = 0
            self.crop_count = 0

            # Creating experiment images
            for image_file_name in self.image_file_names:
                # Getting the image's file path
                self.image_file_path = os.path.join(self.current_folder_path, image_file_name)

                # Assigning an experiment subject ID
                self.exp_subject_id = 'e' + str(self.exp_i - 1)

                # Getting the image's position on the slab and naming its file accordingly
                self.row_changed, self.row, self.col, self.crop_count = self.assign_image_position(image_file_name)
                self.exp_file_name = self.assign_image_name(image_file_name)

                # Renaming, moving the image to the experiment folder, creating a Pillow image instance, and getting exif info
                self.exp_image_file_path = os.path.join(self.exp_folder_path, self.exp_file_name)
                self.exp_image_file_path = shutil.copyfile(self.image_file_path, self.exp_image_file_path)
                self.exp_pil_img, self.image_exif = self.configure_pil_img(self.exp_image_file_path)

                # Getting image's millimeters per pixel ratio
                self.mm_per_pixel = self.get_mm_per_pixel(self.exp_image_file_path)

                # Getting the mean grain size and grain density of the image's granite
                self.grain_density, self.grain_stats = self.get_grain_stats(self.exp_image_file_path)

                # Getting the image's total glare area and a list of its glare pixels
                self.total_glare_area, self.glare_pixels = self.get_glare_area(self.exp_pil_img)

                # Resizing the image to the Zooniverse recommended 600MB and saving changes
                self.exp_pil_img = self.resize_to_limit(self.exp_pil_img, self.exp_image_file_path)

                # TODO: TEMPORARY
                # Drawing scale bars on the image (if specified)
                if self.should_draw_scale_bars == 'y':
                    self.exp_pil_img = self.draw_scale_bars(self.exp_pil_img, self.exp_image_file_path, 10)
                # Resizing the image to the Zooniverse recommended 600MB and saving changes
                self.exp_pil_img = self.resize_to_limit(self.exp_pil_img, self.exp_image_file_path)

                # Getting image's updated (post-resizing) millimeters per pixel ratio
                self.mm_per_pixel = self.get_mm_per_pixel(self.exp_image_file_path)

                # Getting data and GPS info from the image's exif data
                self.exif_dict = {ExifTags.TAGS[k]: j for k, j in self.exp_pil_img._getexif().items() if
                                  k in ExifTags.TAGS}
                self.date = self.get_date_exif()
                self.latitude_longitude = self.get_gps_exif()

                # Writing the image's metadata into the experiment manifest and experiment CSV
                self.exp_metadata_list = [self.exp_subject_id, str(self.exp_file_name), self.warehouse_name, self.location,
                                          self.granite_type, self.slab_id,
                                          self.date.replace("\"", "") if self.date is not None else "",
                                          str(self.latitude_longitude[0]) + ", " + str(self.latitude_longitude[1])
                                          if self.latitude_longitude is not None else "",
                                          self.grain_density, self.grain_stats, self.total_glare_area,
                                          self.number_of_columns]
                self.write_metadata_into_manifest(self.exp_ws, self.exp_i, self.exp_metadata_list)
                self.write_metadata_into_CSV(self.exp_csv_file_path, self.exp_metadata_fields, self.exp_metadata_list)

                # Printing status
                print('\r{} of {} images processed.'.format((self.image_file_names.index(image_file_name) + 1),
                                                            len(self.image_file_names)), end="")

                # Updating indexing variables
                self.exp_i += 1
                self.crop_count += 1

            # Saving experiment manifest
            self.exp_wb.save(self.exp_manifest_path)

            # Creating simulation images
            self.image_file_names_sample_sims = random.sample(self.get_file_names(self.exp_folder_path), self.training_images_per_folder)
            for image_file_name in self.image_file_names_sample_sims:
                # Getting the image's file path
                self.image_file_path = os.path.join(self.exp_folder_path, image_file_name)

                # Assigning a simulation subject ID
                self.sim_subject_id = 's' + str(self.sim_i - 1)

                # Renaming, moving the image to the simulation folder, creating a Pillow image instance, and getting exif info
                self.sim_image_file_name = self.sim_subject_id + r"_" + image_file_name
                self.sim_image_file_path = os.path.join(self.sim_folder_path, self.sim_image_file_name)
                self.sim_image_file_path = shutil.copyfile(self.image_file_path, self.sim_image_file_path)
                self.sim_pil_img, self.image_exif = self.configure_pil_img(self.sim_image_file_path)

                # Drawing the simulation
                self.sim_pil_img, self.ellipse_axes_lengths, self.ellipse_center_coordinates, self.ellipse_angle, \
                    self.major_to_minor_axes_ratio = self.draw_sim(self.sim_pil_img)

                # TODO: TEMPORARY
                # If scale bars were drawn, redrawing them to ensure than they were not covered by a simulation
                if self.should_draw_scale_bars == 'y':
                    self.sim_pil_img = self.draw_scale_bars(self.sim_pil_img, self.sim_image_file_path, 8)

                # Resizing the image to the Zooniverse recommended 600MB and saving changes
                self.sim_pil_img = self.resize_to_limit(self.sim_pil_img, self.sim_image_file_path)

                # Writing the image's metadata into the simulation manifest and simulation CSV
                self.sim_metadata_list = [self.sim_subject_id, self.sim_image_file_name, self.sim_feedback_id,
                                          self.ellipse_center_coordinates[0], self.ellipse_center_coordinates[1],
                                          int(self.ellipse_axes_lengths[0]/self.mm_per_pixel),
                                          int(self.ellipse_axes_lengths[1]/self.mm_per_pixel),
                                          self.ellipse_angle, self.major_to_minor_axes_ratio]
                self.write_metadata_into_manifest(self.sim_ws, self.sim_i, self.sim_metadata_list)
                self.write_metadata_into_CSV(self.sim_csv_file_path, self.sim_metadata_fields, self.sim_metadata_list)

                if self.image_file_names_sample_sims.index(image_file_name) == 0:
                    print("")

                # Printing status
                print('\r{} of {} simulations made.'.format(
                    (self.image_file_names_sample_sims.index(image_file_name) + 1),
                    len(self.image_file_names_sample_sims)), end="")

                # Updating indexing variable
                self.sim_i += 1

            # Saving simulation manifest
            self.sim_wb.save(self.sim_manifest_path)

            # Creating negative images
            self.image_file_names_sample_negs = random.sample(self.get_file_names(self.exp_folder_path), self.training_images_per_folder)
            for image_file_name in self.image_file_names_sample_negs:
                # Getting the image's file path
                self.image_file_path = os.path.join(self.exp_folder_path, image_file_name)

                # Assigning a negative subject ID
                self.neg_subject_id = 'n' + str(self.neg_i - 1)

                # Renaming, moving the image to the simulation folder, creating a Pillow image instance, and getting exif info
                self.neg_image_file_name = self.neg_subject_id + r"_" + image_file_name
                self.neg_image_file_path = os.path.join(self.neg_folder_path, self.neg_image_file_name)
                self.neg_image_file_path = shutil.copyfile(self.image_file_path, self.neg_image_file_path)
                self.neg_pil_img, self.image_exif = self.configure_pil_img(self.neg_image_file_path)

                # Getting input for whether the image is "negative" (contains no candidate melt patches)
                self.neg_pil_img, self.contains_melt_patch, self.classification = self.create_negative(image_file_name,
                                                                                                       self.neg_pil_img)
                if self.contains_melt_patch == 'y':
                    continue

                # Resizing the image to the Zooniverse recommended 600MB and saving changes
                self.neg_pil_img = self.resize_to_limit(self.neg_pil_img, self.neg_image_file_path)

                # Writing the image's metadata into the negative manifest and negative CSV
                self.neg_metadata_list = [self.neg_subject_id, self.neg_image_file_name, "True", self.neg_feedback_id,
                                          self.classification]
                self.write_metadata_into_manifest(self.neg_ws, self.neg_i, self.neg_metadata_list)
                self.write_metadata_into_CSV(self.neg_csv_file_path, self.neg_metadata_fields, self.neg_metadata_list)

                # Updating indexing variable
                self.neg_i += 1

            # Saving negative manifest
            self.neg_wb.save(self.neg_manifest_path)

            # Pushing updated manifests to Github
            self.push_manifests()

            # If specified, uploading images
            if self.upload_now == 'y':
                self.upload_images()

    def push_manifests(self):
        # Pushing the updated local manifests to Github
        repo_directory = '.'  # ie. current directory
        repo = Repo(repo_directory)
        files_to_push = [r"manifests/Experiment_Manifest.xlsx", r"manifests/Simulation_Manifest.xlsx",
                         r"manifests/Negative_Manifest.xlsx"]
        commit_message = f"update manifests, date: {date.today()}, last lines (exp, sim, neg): {self.exp_i, self.sim_i, self.neg_i}"
        repo.index.add(files_to_push)
        repo.index.commit(commit_message)
        origin = repo.remote('origin')
        origin.push()
        print('\nManifests pushed.')

    def upload_images(self):
        exp_upload_cmd = 'panoptes subject-set upload-subjects {} {}'.format(self.experiment_set_id,
                                                                             self.exp_csv_file_path)
        sim_upload_cmd = 'panoptes subject-set upload-subjects {} {}'.format(self.simulation_set_id,
                                                                             self.sim_csv_file_path)
        neg_upload_cmd = 'panoptes subject-set upload-subjects {} {}'.format(self.negative_set_id,
                                                                             self.neg_csv_file_path)
        print("")
        try:
            os.system(exp_upload_cmd)
            print('Experiment subjects uploaded.')
        except:
            print('Error uploading EXPERIMENT subjects; upload manually.')
        try:
            os.system(sim_upload_cmd)
            print('Simulation subjects uploaded.')
        except:
            print('Error uploading SIMULATION subjects; upload manually.')
        try:
            os.system(neg_upload_cmd)
            print('Negative subjects uploaded.')
        except:
            print('Error uploading NEGATIVE subjects; upload manually.')

    def create_negative(self, image_file_name, pil_img):
        # Printing status
        if self.image_file_names_sample_negs.index(image_file_name) == 0:
            print("\n\nMaking negative images...")
        print('Displaying image {} of {}'.format(self.image_file_names_sample_negs.index(image_file_name) + 1,
                                                 len(self.image_file_names_sample_negs)))
        # Displaying the sampled image
        plt.imshow(pil_img)
        plt.show(block=False)
        plt.waitforbuttonpress(0)
        plt.close()
        # Getting input for whether the image is "negative" (contains no candidate melt patches)
        contains_melt_patch = input('Does this image contain a potential melt patch? [y/n]: ')
        while contains_melt_patch != 'y' and contains_melt_patch != 'n':
            print('Please enter \'y\' or \'no\'... ')
            contains_melt_patch = input('Does this image contain a potential melt patch? [y/n]: ')
        # Naming, saving the image according to the above input
        if contains_melt_patch == 'n':
            classification = 'negative'
            pil_img.save(self.neg_image_file_path)
        else:
            classification = 'positive'
            pos_file_name = r"pos_" + image_file_name
            pos_file_path = os.path.join(self.pos_folder_path, pos_file_name)
            pil_img.save(pos_file_path)
            print('\nThe image has been saved to the "positives folder."\n')

        return pil_img, contains_melt_patch, classification

    @staticmethod
    def create_blank_pil_img(model_pil_img):
        # Creating a blank (white) image of equal dimensions to model_pil_img, saving to blank_img_path
        blank_img_path = r"sim_tools\blank.png"
        blank_img_np = 255 * np.ones(np.array(model_pil_img).shape, np.uint8)
        cv2.imwrite(blank_img_path, blank_img_np)
        # Creating a PIL instance for the blank image
        blank_pil_img = Image.open(blank_img_path)

        return blank_pil_img

    def get_glare_ellipse_overlap(self, ellipse_pil_img, ellipse_center_coordinates, blank_pil_img):
        # Pasting the simulation ellipse onto a blank (white) image to obtain a list of ellipse pixels
        blank_pil_img.paste(ellipse_pil_img, ellipse_center_coordinates, ellipse_pil_img)
        ellipse_img = np.array(blank_pil_img)
        ellipse_rows, ellipse_cols, _ = np.where(ellipse_img != 255)
        ellipse_pixels = [(row, col) for (row, col) in zip(ellipse_rows, ellipse_cols)]
        overlap = set(self.glare_pixels).intersection(ellipse_pixels)

        return overlap

    def draw_sim(self, sim_pil_img):
        # Creating a PIL instance for the simulation ellipse template
        ellipse_pil_img = Image.open(self.sim_template_path)
        # Initializing ellipse axes lengths (mm) , and angle (in degrees) rotated above the horizontal
        # TODO: select according to distribution
        minor_axis_mm = 3
        # TODO: select according to distribution
        major_to_minor_axes_ratio = 2
        major_axis_mm = minor_axis_mm * major_to_minor_axes_ratio
        ellipse_axes_lengths = (major_axis_mm, minor_axis_mm)
        # Initializing (random) ellipse position, excluding 1mm region around the image's edges
        ellipse_center_x_coord = random.randint(int(1 / self.mm_per_pixel),
                                                int((self.image_dimensions_mm[1] - 1) / self.mm_per_pixel))
        ellipse_center_y_coord = random.randint(int(1 / self.mm_per_pixel),
                                                int((self.image_dimensions_mm[0] - 1) / self.mm_per_pixel))
        ellipse_center_coordinates = (ellipse_center_x_coord, ellipse_center_y_coord)
        # Initializing (random) ellipse angle (in degrees) above the horizontal
        ellipse_angle = random.randint(0, 180)
        # Converting axes lengths from millimeters to pixels, accounting for the template image's blank-space borders
        major_axis_pix = major_axis_mm / self.mm_per_pixel
        minor_axis_pix = minor_axis_mm / self.mm_per_pixel
        major_axis_excess_pix = (28 * (major_axis_pix / 124))
        minor_axis_excess_pix = 28 * (minor_axis_pix / 124)
        adjusted_major_axis_pix = int(major_axis_pix + major_axis_excess_pix)
        adjusted_minor_axis_pix = int(minor_axis_pix + minor_axis_excess_pix)
        adjusted_ellipse_axes_lengths = (adjusted_major_axis_pix, adjusted_minor_axis_pix)
        # Resizing the ellipse
        ellipse_pil_img = ellipse_pil_img.resize(adjusted_ellipse_axes_lengths)  # major, minor
        # Correcting for the errors introduced by blank-space borders / rotation on the ellipse center coordinates
        L, l = ellipse_pil_img.size
        beta = np.arccos(1 / np.sqrt((l ** 2 / L ** 2) + 1))
        d = np.sqrt((1 / 4) * l ** 2 + (1 / 4) * L ** 2)
        if 0 < ellipse_angle <= 90:
            alpha = np.radians(ellipse_angle) - beta
            x1 = d * np.cos(alpha)
            y1 = d * np.sin(alpha)
            b = L * np.sin(np.radians(ellipse_angle))
            x0 = abs(x1)
            y0 = abs(b - y1)
            topLeft_to_center_x = int(ellipse_center_x_coord - x0)
            topLeft_to_center_y = int(ellipse_center_y_coord - y0)
        elif 90 < ellipse_angle <= 180:
            alpha = np.radians(ellipse_angle) + beta
            x1 = d * np.cos(alpha)
            y1 = d * np.sin(alpha)
            u = l * np.sin(np.radians(ellipse_angle - 90))
            x0 = abs(x1)
            y0 = abs(u + y1)
            topLeft_to_center_x = int(ellipse_center_x_coord - x0)
            topLeft_to_center_y = int(ellipse_center_y_coord - y0)
        elif ellipse_angle == 0:
            topLeft_to_center_x = int(ellipse_center_x_coord - (adjusted_major_axis_pix / 2))
            topLeft_to_center_y = int(ellipse_center_y_coord - (adjusted_minor_axis_pix / 2))
        # Rotating the ellipse
        ellipse_pil_img = ellipse_pil_img.rotate(ellipse_angle, expand=True)
        # Ensuring that the ellipse will not be pasted over glare areas
        overlap = 0
        while overlap != set():
            # Pasting the simulation ellipse onto a blank (white) image to obtain a list of ellipse pixels
            blank_pil_img = self.create_blank_pil_img(sim_pil_img)
            overlap = self.get_glare_ellipse_overlap(ellipse_pil_img, ellipse_center_coordinates, blank_pil_img)
        # Pasting the ellipse onto pil_image
        sim_pil_img.paste(ellipse_pil_img, (topLeft_to_center_x, topLeft_to_center_y), ellipse_pil_img)

        return sim_pil_img, ellipse_axes_lengths, ellipse_center_coordinates, ellipse_angle, major_to_minor_axes_ratio

    @staticmethod
    def write_metadata_into_CSV(CSV_file_path, metadata_fields, metadata_list):
        with open(CSV_file_path, 'a', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields)
            row = {}
            for i in range(len(metadata_fields)):
                row[str(metadata_fields[i])] = metadata_list[i]
            csv_writer.writerow(row)

    @staticmethod
    def write_metadata_into_manifest(ws, row, metadata_list):
        for i in range(len(metadata_list)):
            ws.cell(row=row, column=(i + 1)).value = str(metadata_list[i])

    @staticmethod
    def convert_to_degrees(value):
        # Converting GPS exif data to latitude/longitude degree format
        d0 = value[0][0]
        d1 = value[0][1]
        d = float(d0) / float(d1)
        m0 = value[1][0]
        m1 = value[1][1]
        m = float(m0) / float(m1)
        s0 = value[2][0]
        s1 = value[2][1]
        s = float(s0) / float(s1)

        return d + (m / 60.0) + (s / 3600.0)

    def get_gps_exif(self):
        gps_dict = {}
        try:
            for key in self.exif_dict['GPSInfo'].keys():
                gps_tag = ExifTags.GPSTAGS.get(key)
                gps_dict[gps_tag] = self.exif_dict['GPSInfo'][key]
            latitude_raw = gps_dict.get('GPSLatitude')
            longitude_raw = gps_dict.get('GPSLongitude')
            lat_ref = gps_dict.get('GPSLatitudeRef')
            long_ref = gps_dict.get('GPSLongitudeRef')
            if lat_ref == "S":
                latitude = -abs(self.convert_to_degrees(latitude_raw))
            else:
                latitude = self.convert_to_degrees(latitude_raw)
            if long_ref == "W":
                longitude = -abs(self.convert_to_degrees(longitude_raw))
            else:
                longitude = self.convert_to_degrees(longitude_raw)
        except:
            latitude = None
            longitude = None
            pass
        latitude_longitude = [latitude, longitude]

        return latitude_longitude

    def get_date_exif(self):
        # Attempting to retrieve data exif data
        try:
            date_exif = datetime.strptime(self.exif_dict['DateTimeOriginal'], '%Y:%m:%d %H:%M:%S')
            date_exif = json.dumps(date_exif, default=str)
        except:
            date_exif = None
            pass

        return date_exif

    def resize_to_limit(self, pil_img, image_file_path, size_limit=600000):
        aspect = pil_img.size[0] / pil_img.size[1]
        while True:
            with io.BytesIO() as buffer:
                pil_img.save(buffer, format="JPEG")
                data = buffer.getvalue()
            file_size = len(data)
            size_deviation = file_size / size_limit
            if size_deviation <= 1:
                pil_img.save(image_file_path, exif=self.image_exif)
                pil_img = Image.open(image_file_path)
                break
            else:
                new_width = pil_img.size[0] / (1 * (size_deviation ** 0.5))
                new_height = new_width / aspect
                pil_img = pil_img.resize((int(new_width), int(new_height)))

        return pil_img

    def draw_scale_bars(self, pil_img, image_file_path, thickness):
        pix_width, pix_height = pil_img.size
        draw = ImageDraw.Draw(pil_img)
        # Setting scale bar color, length, and number
        scale_bars_color = (100, 255, 0)  # (R, G, B)
        scale_bar_pix_length = self.lower_limit / self.mm_per_pixel
        number_of_scale_bars = 10
        # Configuring left/right side scale bars
        leftRight_bars_center = (pix_height / number_of_scale_bars) / 2
        leftRight_start_coords = [15, leftRight_bars_center - (.5 * scale_bar_pix_length)]
        leftRight_end_coords = [leftRight_start_coords[0], leftRight_start_coords[1] + scale_bar_pix_length]
        leftRight_coords = [tuple(leftRight_start_coords), tuple(leftRight_end_coords)]
        # Configuring top/bottom side scale bars
        topBottom_bars_center = (pix_width / number_of_scale_bars) / 2
        topBottom_start_coords = [topBottom_bars_center - (.5 * scale_bar_pix_length), 15]
        topBottom_end_coords = [topBottom_start_coords[0] + scale_bar_pix_length, topBottom_start_coords[1]]
        topBottom_coords = [tuple(topBottom_start_coords), tuple(topBottom_end_coords)]
        # Drawing scale bars
        for j in range(1, number_of_scale_bars + 1):
            draw.line(leftRight_coords, fill=scale_bars_color, width=thickness)
            draw.line(topBottom_coords, fill=scale_bars_color, width=thickness)
            leftRight_start_coords[1] += (2 * leftRight_bars_center)
            topBottom_start_coords[0] += (2 * topBottom_bars_center)
            if j % 2 != 0:
                leftRight_start_coords[0] = pix_width - leftRight_start_coords[0]
                topBottom_start_coords[1] = pix_height - topBottom_start_coords[1]
            else:
                leftRight_start_coords[0] = 15
                topBottom_start_coords[1] = 15
            leftRight_end_coords = [leftRight_start_coords[0], leftRight_start_coords[1] + scale_bar_pix_length]
            topBottom_end_coords = [topBottom_start_coords[0] + scale_bar_pix_length, topBottom_start_coords[1]]
            leftRight_coords = [tuple(leftRight_start_coords), tuple(leftRight_end_coords)]
            topBottom_coords = [tuple(topBottom_start_coords), tuple(topBottom_end_coords)]
        # Saving image to image_file_path with its exif data
        pil_img.save(image_file_path, exif=self.image_exif)
        pil_img = Image.open(image_file_path)

        return pil_img

    @staticmethod
    def midpoint(ptA, ptB):
        return (ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5

    @staticmethod
    def canny_thresholds(img, sigma=0.33):
        # Compute the median of the single channel pixel intensities
        v = np.median(img)
        # Apply automatic Canny edge detection using the computed median
        lower = int(max(0, (1.0 - sigma) * v))
        upper = int(min(255, (1.0 + sigma) * v))

        return lower, upper

    def get_contours(self, prepared_img, cutoff_avg_dim_mm):
        # Detecting contours
        lower_canny, upper_canny = self.canny_thresholds(prepared_img)
        edges_img = cv2.Canny(prepared_img, lower_canny, upper_canny)
        # Dilating to ensure that contour paths are continuous
        dil_kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        dil_edges_img = cv2.dilate(edges_img, dil_kernel)
        # Getting, sorting a list of contours
        contours_list = cv2.findContours(dil_edges_img.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours_list = grab_contours(contours_list)
        if contours_list:
            (contours_list, _) = contours.sort_contours(contours_list)
        # Initializing lists
        wanted_contours_list = []
        contour_dims = []
        center_xs = []
        center_ys = []
        for cont in contours_list:
            #  Creating bounding box
            box = cv2.minAreaRect(cont)
            box = cv2.boxPoints(box)
            box = np.array(box, dtype='int')
            box = perspective.order_points(box)
            # Getting box vertex coordinates
            (top_left, top_right, bot_right, bot_left) = box
            # Getting box side midpoints
            (top_mid_x, top_mid_y) = self.midpoint(top_left, top_right)
            (bot_mid_x, bot_mid_y) = self.midpoint(bot_left, bot_right)
            (left_mid_x, left_mid_y) = self.midpoint(top_left, bot_left)
            (right_mid_x, right_mid_y) = self.midpoint(top_right, bot_right)
            # Getting box center coordinates
            center_x = np.average(box[:, 0])
            center_y = np.average(box[:, 1])
            center_xs.append(center_x)
            center_ys.append(center_y)
            # Measuring the semi-major and semi-minor axes of the box in pixels
            pix_height = distance.euclidean((top_mid_x, top_mid_y), (bot_mid_x, bot_mid_y))
            pix_width = distance.euclidean((left_mid_x, left_mid_y), (right_mid_x, right_mid_y))
            contour_dim = (pix_height, pix_width)
            # Converting axes lengths to millimeters
            mm_height = pix_height * self.mm_per_pixel
            mm_width = pix_width * self.mm_per_pixel
            # Getting the contour's average dimension (average of height/width) in millimeters
            contour_avg_dim = (mm_height + mm_width) / 2
            # Ignoring contours that have an average dimension less than the specified cutoff
            if contour_avg_dim < cutoff_avg_dim_mm:
                continue
            wanted_contours_list.append(cont)
            contour_dims.append(contour_dim)
        # Getting a list of contour areas in square pixels
        contour_areas = []
        for cont in wanted_contours_list:
            area = cv2.contourArea(cont)
            contour_areas.append(area)
        # Finding the total contour area in square pixels
        total_contour_area = sum(contour_areas)

        return wanted_contours_list, contour_dims, total_contour_area

    def get_glare_area(self, pil_img):
        cv2_img = np.array(pil_img)
        # Setting lower and upper color limits for thresholding
        lower = (230, 230, 230)
        upper = (255, 255, 255)
        # Getting pixels between lower and upper color limits (those corresponding to glare areas), blurring to smooth edges
        mask = cv2.inRange(cv2_img, lower, upper)
        mask = cv2.blur(mask, (10, 10))
        # Getting a list of the glare patches' edges/contours, dimensions and the total glare area
        glare_contours, glare_dims, total_glare_area = self.get_contours(mask, 2.5)
        # Converting the total glare area from square pixels to square millimeters
        total_glare_area_mmSq = total_glare_area * (self.mm_per_pixel ** 2)
        # Isolating glare areas on a blank (black) image to get a list of glare pixels
        blank_img = np.zeros(cv2_img.shape, np.uint8)
        isolated_glare_img = cv2.drawContours(blank_img, glare_contours, contourIdx=-1, color=(255, 255, 255),
                                              thickness=cv2.FILLED)
        glare_rows, glare_cols, _ = np.where(isolated_glare_img != 0)
        glare_pixels = [(row, col) for (row, col) in zip(glare_rows, glare_cols)]

        return total_glare_area_mmSq, glare_pixels

    def get_grain_stats(self, image_file_path):
        """Note: "grains" refers to the dark-colored grains in the granite"""
        original_img = cv2.imread(image_file_path)
        original_img = img_as_ubyte(original_img)
        # Converting to grayscale
        gray_img = cv2.cvtColor(original_img, cv2.COLOR_BGR2GRAY)
        # Blurring grayscale image
        blurred_img = cv2.GaussianBlur(gray_img, (9, 9), 0)
        # Segmenting blurred image to separate dark grains from background
        segmentation_thresh = 70
        thresh_img = cv2.threshold(blurred_img, segmentation_thresh, 255, cv2.THRESH_BINARY)[1]
        # Opening (erosion followed by dilation) segmented image to close holes within grains
        kernel = np.ones((9, 9), np.uint8)
        opened_img = cv2.morphologyEx(thresh_img, cv2.MORPH_OPEN, kernel, 5)
        # Getting a list of the grains' edges/contours, dimensions, and total area
        grain_contours_list, grain_dims, total_grain_area = self.get_contours(opened_img, 0.015)
        # Converting the grain dimensions from pixels to millimeters
        grain_dims_mm = [(x * self.mm_per_pixel, y * self.mm_per_pixel) for x, y in grain_dims]
        # Finding the total image area in square pixels
        image_area = (self.image_dimensions[0] * self.image_dimensions[1]) * (self.mm_per_inch ** 2) / (
                self.mm_per_pixel ** 2)
        # Getting the image's grain density (area of grains / area of image)
        grain_density = total_grain_area / image_area
        # Getting statistics on the image's grains
        h = [g[0] for g in grain_dims_mm]
        w = [g[1] for g in grain_dims_mm]
        number_of_grains = len(grain_dims_mm)
        mean_grain_size = statistics.mean([statistics.mean(h), statistics.mean(w)])
        median_grain_size = statistics.median([statistics.median(h), statistics.median(w)])
        grain_size_25_percentile = statistics.mean([np.percentile(h, 25), np.percentile(w, 25)])
        grain_size_75_percentile = statistics.mean([np.percentile(h, 75), np.percentile(w, 75)])
        grain_stats = (
            number_of_grains, mean_grain_size, median_grain_size, grain_size_25_percentile, grain_size_75_percentile)

        return grain_density, grain_stats

    def get_mm_per_pixel(self, image_file_path):
        cv2_img = cv2.imread(image_file_path)
        pix_height, pix_width, _ = cv2_img.shape
        mm_per_pixel = self.image_dimensions_mm[0] / pix_height

        return mm_per_pixel

    def assign_image_name(self, image_file_name):
        extension = os.path.splitext(image_file_name)[-1]
        if self.should_crop_into_four == 'y':
            crop_file_name = os.path.splitext(image_file_name)[-2]
            crop_location = crop_file_name.split("_")[-1]
            exp_file_name = self.exp_subject_id + r"_" + str(self.slab_id) + r"_" + str(
                self.row) + r"_" + str(self.col) + r"_" + \
                            crop_location + str(extension)
        else:
            exp_file_name = self.exp_subject_id + r"_" + str(self.slab_id) + r"_" + str(self.row) + r"_" + str(
                self.col) + str(
                extension)

        return exp_file_name

    def assign_image_position(self, image_file_name):
        row_changed = 0
        if self.should_crop_into_four == 'y' and self.number_of_columns:
            if (self.image_file_names.index(image_file_name) / 4) % int(
                    self.number_of_columns) == 0 and self.image_file_names.index(image_file_name) != 0:
                self.row += 1
                self.row_changed = 1
        elif self.should_crop_into_four == 'n':
            if self.image_file_names.index(image_file_name) % int(
                    self.number_of_columns) == 0 and self.image_file_names.index(image_file_name) != 0:
                self.row += 1
                self.row_changed = 1
        # Getting column number
        if self.row % 2 == 0:
            if self.should_crop_into_four == 'y':
                if self.crop_count % 4 == 0 and row_changed != 1:
                    self.col -= 1
            elif self.should_crop_into_four == 'n' and row_changed != 1:
                self.col -= 1
        elif self.row % 2 != 0:
            if self.should_crop_into_four == 'y' and row_changed != 1:
                if self.crop_count % 4 == 0:
                    self.col += 1
            elif self.should_crop_into_four == 'n' and row_changed != 1:
                self.col += 1
        else:
            self.row = 0

        return row_changed, self.row, self.col, self.crop_count

    @staticmethod
    def configure_pil_img(image_file_path):
        pil_img = Image.open(image_file_path)
        image_exif = pil_img.info['exif']

        return pil_img, image_exif

    def crop_into_four(self):
        for file_name in self.image_file_names:
            image_file_path = os.path.join(self.current_folder_path, file_name)
            cropped_file_path = os.path.join(self.cropped_folder_path, file_name)
            extension = os.path.splitext(file_name)[-1]
            pil_img = Image.open(image_file_path)
            image_exif = pil_img.info['exif']
            width, height = pil_img.size
            half_width = width / 2
            half_height = height / 2
            # Starting from top left (0,0) and moving clockwise
            section_1 = (0, 0, half_width, half_height)  # (Left, Top, Right, Bottom) Starting pts.
            section_2 = (half_width, 0, width, half_height)
            section_3 = (half_width, half_height, width, height)
            section_4 = (0, half_height, half_width, height)
            location = ''
            for j in range(1, 5):
                if j == 1:
                    pil_img = pil_img.crop(section_1)
                    location = 'TL'
                if j == 2:
                    pil_img = pil_img.crop(section_2)
                    location = 'TR'
                if j == 3:
                    pil_img = pil_img.crop(section_3)
                    location = 'BR'
                if j == 4:
                    pil_img = pil_img.crop(section_4)
                    location = 'BL'
                reformatted_cropped_file_path = cropped_file_path.replace(str(extension),
                                                                          "_{}{}".format(location, str(extension)))
                pil_img.save(reformatted_cropped_file_path, exif=image_exif)
        self.image_file_names = self.get_file_names(self.cropped_folder_path)
        self.current_folder_path = self.cropped_folder_path
        print('\nImages cropped.\n')

        return self.current_folder_path, self.image_file_names

    @staticmethod
    def get_file_names(folder_path):
        all_file_names = os.listdir(folder_path)
        wanted_file_names = []
        for file_name in all_file_names:
            if file_name.endswith(".jpeg") or file_name.endswith(".jpg") or file_name.endswith(".png"):
                wanted_file_names.append(file_name)

        return wanted_file_names

    @staticmethod
    def configure_exp_metadata():
        warehouse_name = input('Warehouse name: ')
        location = input('Location (City, State): ')
        granite_type = input('Granite type: ')
        slab_id = input('Slab ID: ')
        number_of_columns = input('Number of Columns: ')
        print("")

        # warehouse_name = 'United Stone International'
        # location = 'Solon, Ohio'
        # granite_type = 'Dallas White'
        # slab_id = '1151|20'

        return warehouse_name, location, granite_type, slab_id, number_of_columns

    @staticmethod
    def configure_CSV(subfolder_path, desired_csv_file_name, metadata_fields_list=None):
        if metadata_fields_list is None:
            metadata_fields_list = ['']
        desired_csv_file_path = os.path.join(subfolder_path, desired_csv_file_name)
        with open(desired_csv_file_path, 'w', newline='') as f:
            csv_writer = csv.DictWriter(f, fieldnames=metadata_fields_list)
            csv_writer.writeheader()

        return desired_csv_file_path

    def configure_CSVs(self):
        exp_csv_file_name = 'experiment_subjects.csv'
        exp_metadata_fields_list = ['!subject_id', '#file_name', '#warehouse', '#location', '#granite_type', '#slab_id',
                                    '#date_time', '#latitude_longitude', '#grain_density',
                                    '#grain_stats(mm)(number_of_grains, '
                                    'mean_grain_size, median_grain_size, grain_size_25_percentile, grain_size_75_percentile)',
                                    '#glare_area(mm^2)', '#number_of_columns']
        exp_csv_file_path = self.configure_CSV(self.exp_folder_path, exp_csv_file_name, exp_metadata_fields_list)

        sim_csv_file_name = 'simulation_subjects.csv'
        sim_metadata_fields_list = ['!subject_id', '#file_name', '#feedback_1_id', '#feedback_1_x',
                                    '#feedback_1_y', '#feedback_1_toleranceA', '#feedback_1_toleranceB',
                                    '#feedback_1_theta',
                                    '#minor_to_major_ratio']
        sim_csv_file_path = self.configure_CSV(self.sim_folder_path, sim_csv_file_name, sim_metadata_fields_list)

        neg_csv_file_name = 'negative_subjects.csv'
        neg_metadata_fields_list = ['!subject_id', '#file_name', '#training_subject', '#feedback_1_id',
                                    '#classification']
        neg_csv_file_path = self.configure_CSV(self.neg_folder_path, neg_csv_file_name, neg_metadata_fields_list)

        return exp_csv_file_name, exp_csv_file_path, exp_metadata_fields_list, sim_csv_file_name, sim_csv_file_path, \
            sim_metadata_fields_list, neg_csv_file_name, neg_csv_file_path, neg_metadata_fields_list

    @staticmethod
    def configure_excel(excel_file_path):
        wb = openpyxl.load_workbook(filename=excel_file_path)
        ws = wb['Sheet1']
        first_empty_row = None
        for row in range(1, int(1e10)):
            if ws.cell(row, 1).value is None:
                first_empty_row = row
                break

        return wb, ws, first_empty_row

    @staticmethod
    def clear_folder(folder_path):
        try:
            shutil.rmtree(folder_path)
            os.mkdir(folder_path)
        except PermissionError or FileNotFoundError:
            input('\nPermissions error for {}, '
                  '\nExit the file window and restart the program.'.format(folder_path))

    def make_folder(self, current_folder_path, desired_folder_name=""):
        desired_folder_path = os.path.join(current_folder_path, desired_folder_name)
        try:
            os.mkdir(desired_folder_path)
        except FileExistsError:
            self.clear_folder(desired_folder_path)

        return desired_folder_path

    def make_folders(self):
        experiment_folder_path = self.make_folder(self.current_folder_path, r"experiment")
        if self.should_crop_into_four == 'y':
            cropped_folder_path = self.make_folder(self.current_folder_path, r"cropped")
        else:
            cropped_folder_path = ''
        sim_folder_path = self.make_folder(self.current_folder_path, r"simulations")
        neg_folder_path = self.make_folder(self.current_folder_path, r"negatives")
        pos_folder_path = self.make_folder(self.current_folder_path, r"positives")

        return experiment_folder_path, cropped_folder_path, sim_folder_path, neg_folder_path, pos_folder_path

    def configure_designator(self):
        # Connect to project & workflow
        zoonv_project, zoonv_workflow = self.configure_zooniverse()

        # Configuring training sets, training image probabilities
        zoonv_workflow.configuration['training_set_ids'] = [self.simulation_set_id, self.negative_set_id]
        zoonv_workflow.configuration['training_chances'] = self.training_chances
        zoonv_workflow.configuration['training_default_chances'] = self.training_default
        zoonv_workflow.configuration[
            'subject_queue_page_size'] = 10  # determines how many subjects are loaded in queue at one time

        # TODO: CHANGE to 15?
        # Training subjects are not retired, experiment subjects are retired via SWAP/Caesar
        zoonv_workflow.retirement['criteria'] = 'never_retire'

        # Saving
        zoonv_workflow.modified_attributes.add('configuration')
        zoonv_workflow.save()

        print('\nDesignator configured.')

    def get_existing_subject_sets(self):
        # Connecting to project & workflow
        zoonv_project, zoonv_workflow = self.configure_zooniverse()

        # Getting and printing a list of existing subject sets' names & IDs
        subject_set_ids = []
        subject_set_names = []
        print('\nThe existing subject sets are:')
        for ss in zoonv_project.links.subject_sets:
            subject_set_name = ss.display_name
            subject_set_id = int(str(ss).split()[1].replace('>', ''))
            subject_set_ids.append(subject_set_id)
            subject_set_names.append(subject_set_name)
        print(
            "\n".join(u"{} \u2014 {}".format(ss_id, ss_name) for ss_id, ss_name in zip(subject_set_ids, subject_set_names)))
        return subject_set_ids

    def configure_subject_set(self, subject_type, subject_set_ids):
        # Connecting to project & workflow
        zoonv_project, zoonv_workflow = self.configure_zooniverse()

        # Selecting an existing subject set or creating a new one based on user input
        if subject_type == 'experiment':
            print("")
        need_new_set = input(f"Would you like to create a new {subject_type.upper()} subject set? [y/n]: ")
        while need_new_set != 'y' and need_new_set != 'n':
            print("Please enter \'y\' or \'no\'...")
            need_new_set = input(f"\nWould you like to create a new {subject_type.upper()} subject set? [y/n]: ")
        if need_new_set == 'n':
            subject_set_id = input("    Enter the ID of the existing set you\'d like to upload to: ")
            while (int(subject_set_id) in subject_set_ids) is False:
                subject_set_id = input("    This ID does not exist; please enter a new one: ")
        else:
            subject_set_name = input("    Enter a name for the new subject set: ")
            subject_set = SubjectSet()
            subject_set.links.project = zoonv_project
            subject_set.display_name = subject_set_name
            subject_set.save()
            zoonv_workflow.links.subject_sets.add(subject_set)
            zoonv_workflow.save()
            subject_set_id = int(str(subject_set).split()[1].replace('>', ''))

        return subject_set_id, need_new_set

    def configure_zooniverse(self):
        # Connect to Zooniverse, get project & workflow ID's
        Panoptes.connect(username=self.username, password=self.password)
        zoonv_project = Project.find(self.project_id)
        zoonv_project.save()
        zoonv_workflow = Workflow.find(self.workflow_id)
        zoonv_workflow.save()

        return zoonv_project, zoonv_workflow
